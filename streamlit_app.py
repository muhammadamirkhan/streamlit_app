import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import os
import json
import datetime as _dtm
import requests
from io import BytesIO

st.set_page_config(page_title="Muraba Veil – Unit Manager", layout="wide", page_icon="🏙️")

# ── Password gate ──────────────────────────────────────────────────────────────
def _check_password() -> bool:
    expected = st.secrets.get("password", os.environ.get("APP_PASSWORD", "muraba2026"))

    def _entered():
        st.session_state["auth_ok"] = st.session_state.get("pw_input", "") == expected
        st.session_state.pop("pw_input", None)

    if st.session_state.get("auth_ok"):
        return True

    st.markdown("## 🔒 Muraba Veil — Unit Manager")
    st.text_input("Enter password to continue", type="password",
                  on_change=_entered, key="pw_input")
    if st.session_state.get("auth_ok") is False:
        st.error("Incorrect password — try again.")
    st.caption("Access is restricted. Contact the administrator for the password.")
    return False

if not _check_password():
    st.stop()

# ── Data file (lives next to this script, so it works locally and on the cloud) ─
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Muraba Veil Unit list.xlsx")
COMMENTS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "muraba_comments.json")
STATE_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_state.json")
BASE_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "base_version.json")


# ── Saved working state (explicit Save / Reset) ────────────────────────────────

def _state_dict():
    """The full working state (register, floors, params, MEP map) as a plain dict."""
    return {
        "units": json.loads(st.session_state.units.to_json(orient="records")),
        "floors": st.session_state.floors,
        "params": st.session_state.fm_params,
        "uid_counter": int(st.session_state.get("uid_counter", len(st.session_state.units))),
        # MEP / Majlis floors can be renumbered (MEP-moves), so persist them with the state
        "blocked": {str(k): v for k, v in st.session_state.get("blocked", {}).items()},
    }

def _write_state(path):
    """Snapshot the full working state to a JSON file."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(_state_dict(), f, ensure_ascii=False, indent=2)

def state_json_bytes():
    """The current working state as downloadable JSON bytes."""
    return json.dumps(_state_dict(), ensure_ascii=False, indent=2).encode("utf-8")

def save_state():
    """Persist the current working state (shown on every launch)."""
    _write_state(STATE_PATH)

def save_base():
    """Persist the current working state as the separate Base Version snapshot."""
    _write_state(BASE_PATH)

def _parse_state(state):
    """Turn a state dict into (units_df, floors, params, uid_counter, blocked)."""
    units = pd.DataFrame(state["units"])
    # backfill columns older snapshots may not have, so all downstream code is safe
    for col in ("Dup_Up", "Terrace_Override", "Sellable_Override", "Adj_Pct"):
        if col not in units.columns:
            units[col] = pd.NA
    if "Comment" not in units.columns:
        units["Comment"] = ""
    if "uid" not in units.columns:
        units["uid"] = [f"u{i}" for i in range(len(units))]
    for ovc in ("Terrace_Override", "Sellable_Override"):
        units[ovc] = units[ovc].where(units[ovc].notna(), pd.NA)
    units["Comment"] = units["Comment"].fillna("").astype(str)
    blk = state.get("blocked")
    blk = {int(k): v for k, v in blk.items()} if blk else None
    return units, state["floors"], state["params"], int(state.get("uid_counter", len(units))), blk

def load_state_from(path):
    """Return (units_df, floors, params, uid_counter, blocked) from a state file, or None on failure.
    `blocked` is the persisted MEP/Majlis map (or None if the file predates MEP-moves)."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return _parse_state(json.load(f))
    except Exception:
        return None

def load_state():
    return load_state_from(STATE_PATH)

def load_base():
    return load_state_from(BASE_PATH)

def base_unit_count():
    """Number of units in the saved Base Version (for the 'Additional from base' card), or None."""
    try:
        with open(BASE_PATH, "r", encoding="utf-8") as f:
            return len(json.load(f).get("units", []))
    except Exception:
        return None

def has_saved_state():
    return os.path.exists(STATE_PATH)

def has_base():
    return os.path.exists(BASE_PATH)

def clear_saved_state():
    try:
        os.remove(STATE_PATH)
    except Exception:
        pass


def comment_key(unit, type_, floor):
    return f"{unit}|{type_}|{floor}"

def load_comments_file() -> dict:
    try:
        with open(COMMENTS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_comments_file(mapping: dict):
    try:
        with open(COMMENTS_PATH, "w", encoding="utf-8") as f:
            json.dump(mapping, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def persist_all_comments():
    """Write every non-empty comment in the register to the JSON file, keyed by unit/type/floor."""
    u = st.session_state.units
    mapping = {}
    for _, r in u.iterrows():
        c = r.get("Comment", "")
        if isinstance(c, str) and c.strip():
            mapping[comment_key(r["Unit"], r["Type"], r["Floor"])] = c
    save_comments_file(mapping)


# ── Named versions store (Supabase) ────────────────────────────────────────────
# A small library of full-state snapshots, each addressable by name, kept in a Supabase
# table so they survive Cloud redeploys. We talk to the PostgREST endpoint directly with
# `requests` (no heavy SDK) and only ever call out on explicit Save / Load / list actions —
# the list is cached — so the normal in-memory hot path is never touched.
SUPABASE_TABLE = "muraba_versions"

def _sb_cfg():
    """(base_url, api_key) from secrets/env, both stripped; ('', '') when not configured."""
    try:
        url = st.secrets.get("supabase_url", os.environ.get("SUPABASE_URL", ""))
        key = st.secrets.get("supabase_key", os.environ.get("SUPABASE_KEY", ""))
    except Exception:
        url, key = os.environ.get("SUPABASE_URL", ""), os.environ.get("SUPABASE_KEY", "")
    return (url or "").rstrip("/"), (key or "")

def sb_enabled():
    u, k = _sb_cfg()
    return bool(u and k)

def _sb_headers(key):
    return {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

def _sb_endpoint():
    u, _ = _sb_cfg()
    return f"{u}/rest/v1/{SUPABASE_TABLE}"

@st.cache_data(ttl=300, show_spinner=False)
def sb_list_versions(_nonce=0):
    """[(name, updated_at), …] newest first. Cached; pass a changing _nonce to force a refresh."""
    u, k = _sb_cfg()
    if not (u and k):
        return []
    try:
        r = requests.get(_sb_endpoint(), headers=_sb_headers(k),
                         params={"select": "name,updated_at", "order": "updated_at.desc"}, timeout=10)
        r.raise_for_status()
        return [(d["name"], d.get("updated_at")) for d in r.json()]
    except Exception:
        return []

def sb_save_version(name, state):
    """Upsert a named snapshot (overwrites if the name already exists)."""
    u, k = _sb_cfg()
    if not (u and k):
        raise RuntimeError("Supabase is not configured.")
    body = {"name": name, "state": state, "updated_at": _dtm.datetime.now(_dtm.timezone.utc).isoformat()}
    h = _sb_headers(k); h["Prefer"] = "resolution=merge-duplicates,return=minimal"
    r = requests.post(_sb_endpoint(), headers=h, params={"on_conflict": "name"},
                      data=json.dumps(body), timeout=15)
    r.raise_for_status()

def sb_load_version(name):
    """The stored state dict for `name`, or None."""
    u, k = _sb_cfg()
    if not (u and k):
        return None
    r = requests.get(_sb_endpoint(), headers=_sb_headers(k),
                     params={"select": "state", "name": f"eq.{name}", "limit": "1"}, timeout=15)
    r.raise_for_status()
    data = r.json()
    return data[0]["state"] if data else None

def sb_delete_version(name):
    u, k = _sb_cfg()
    if not (u and k):
        return
    r = requests.delete(_sb_endpoint(), headers=_sb_headers(k),
                        params={"name": f"eq.{name}"}, timeout=15)
    r.raise_for_status()

UNIT_TYPES = [
    "2 Bedroom", "3 Bedroom - New", "3 Bedroom", "3 Bedroom Pool", "4 Bedroom Pool",
    "4 Bedroom XL", "3 Bedroom Duplex", "4 Bedroom Duplex", "5 Bedroom Duplex",
]
STATUS_OPTIONS = ["Available", "Sold"]

# "3 Bedroom - New" is a 2-Bedroom-sized unit under a new name; it shares the
# 2 Bedroom price ladder, area, terrace and escalation (same stats).
PRICE_FAMILY = {"3 Bedroom - New": "2 Bedroom"}

def family_types(t):
    """All types that share a price ladder with t (e.g. 3 Bedroom - New ↔ 2 Bedroom)."""
    base = PRICE_FAMILY.get(t, t)
    fam = {t, base}
    fam.update(k for k, v in PRICE_FAMILY.items() if v == base)
    return fam

TYPE_DEFAULTS = {
    "2 Bedroom":         {"internal": 2218.764851, "external": 1619.322681, "parking": 2, "terrace_rate": 0.30, "levels": 1},
    "3 Bedroom - New":   {"internal": 2218.764851, "external": 1619.322681, "parking": 2, "terrace_rate": 0.30, "levels": 1},
    "3 Bedroom":         {"internal": 2880.530062, "external": 2058.920781, "parking": 2, "terrace_rate": 0.30, "levels": 1},
    "3 Bedroom Pool":    {"internal": 2880.530062, "external": 2059.243699, "parking": 2, "terrace_rate": 0.65, "levels": 2},
    "4 Bedroom Pool":    {"internal": 4643.550947, "external": 5258.816065, "parking": 3, "terrace_rate": 0.55, "levels": 2},
    "4 Bedroom XL": {"internal": 7474.889938, "external": 7857.654592, "parking": 4, "terrace_rate": 0.65, "levels": 1},
    "3 Bedroom Duplex":  {"internal": 4733.537238, "external": 3334.228886, "parking": 3, "terrace_rate": 0.75, "levels": 2},
    "4 Bedroom Duplex":  {"internal": 7485.653849, "external": 7260.042287, "parking": 4, "terrace_rate": 0.75, "levels": 2},
    "5 Bedroom Duplex":  {"internal": 11648.17,    "external": 15018.56,    "parking": 6, "terrace_rate": 1.00, "levels": 2},
}

LEVEL_CAPACITY = {"2 Bedroom": 2, "3 Bedroom": 1}   # standard residential floor

TYPE_ABBR = {
    "2 Bedroom": "2BR", "3 Bedroom - New": "3BR New", "3 Bedroom": "3BR",
    "3 Bedroom Pool": "3BR Pool", "4 Bedroom Pool": "4BR Pool", "4 Bedroom XL": "4BR XL",
    "3 Bedroom Duplex": "3BR DX", "4 Bedroom Duplex": "4BR DX", "5 Bedroom Duplex": "5BR DX",
}

# Fallback escalation defaults (overridden by what we read from the sheets)
ESC_DEFAULTS = {
    "2 Bedroom": 150.0, "3 Bedroom - New": 150.0, "3 Bedroom": 150.0,
    "3 Bedroom Pool": 104.0, "4 Bedroom Pool": 104.0,
    "4 Bedroom XL": 497.0, "3 Bedroom Duplex": 308.0,
    "4 Bedroom Duplex": 305.0, "5 Bedroom Duplex": 0.0,
}
# Terrace-rate groups (% of internal), variable; from Launches XL SX DX
TERRACE_DEFAULTS = {
    "standard": 0.30,        # 2BR & 3BR
    "3 Bedroom Pool": 0.65,  # 3 Pool Terrace
    "4 Bedroom Pool": 0.55,  # 4 Pool Terrace Rate
    "duplex": 0.75,          # DX Terrace Rate (3/4/5 BR Duplex)
    "simplex": 0.65,         # SX Terrace Rate (4 BR Simplex / XL)
}

BLUE_DARK = "#1F4E78"
BLUE_MED  = "#2E75B6"
BLUE_LITE = "#DDEBF7"

SQFT_PER_SQM = 10.7639   # 1 m² = 10.7639 ft²  →  sqm = sqft / 10.7639


# ── Excel-style table renderer (first column left, numeric columns centered) ───

TOTAL_LABELS = {"total", "grand total", "totals"}

def _is_total_row(first_cell):
    return str(first_cell).strip().lower() in TOTAL_LABELS

def excel_table(df: pd.DataFrame):
    def _hl_total(row):
        if _is_total_row(row.iloc[0]):
            return [f"font-weight:bold;background-color:{BLUE_MED};color:#FFFFFF;" for _ in row]
        return ["" for _ in row]
    sty = (df.style.hide(axis="index").apply(_hl_total, axis=1).set_table_styles([
        {"selector": "", "props": "border-collapse:collapse;font-size:13px;width:100%;"
                                   "font-family:Calibri,Arial,sans-serif;"},
        {"selector": "thead th", "props": f"background-color:{BLUE_DARK};color:#FFFFFF;font-weight:bold;"
                                           "text-align:center;border:1px solid #9DC3E6;padding:6px 10px;"},
        {"selector": "tbody td", "props": "border:1px solid #BDD7EE;padding:5px 10px;text-align:center;white-space:nowrap;"},
        {"selector": "tbody td:first-child", "props": "text-align:left;font-weight:600;white-space:nowrap;"},
        {"selector": "thead th:first-child", "props": "text-align:left;white-space:nowrap;"},
        {"selector": "tbody tr:nth-child(even)", "props": f"background-color:{BLUE_LITE};"},
        {"selector": "tbody tr:nth-child(odd)",  "props": "background-color:#FFFFFF;"},
    ]))
    st.markdown(f'<div style="overflow-x:auto">{sty.to_html()}</div>', unsafe_allow_html=True)


def df_to_styled_xlsx(df: pd.DataFrame, sheet_name="Sheet1", title=None, aed_cols=None, sold_mask=None):
    """One formatted sheet matching the client's Excel (Blue Accent-1 theme):
    medium-blue header, banded data rows, bold Total row, dark-blue text. Optionally
    formats `aed_cols` as AED currency and paints Sold rows (per `sold_mask`) yellow."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    # Office "Blue, Accent 1" palette
    HEAD, DATA, TOTAL = "4472C4", "D9E1F2", "B4C6E7"
    DARK, WHITE, SOLD = "1F3864", "FFFFFF", "FFFF00"
    aed_cols = set(aed_cols or [])
    aed_idx = {i + 1 for i, c in enumerate(df.columns) if c in aed_cols}     # 1-based Excel cols
    sold_mask = list(sold_mask) if sold_mask is not None else None
    sheet_name = (sheet_name or "Sheet1")[:31]
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        startrow = 1 if title else 0
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)
        ws = writer.book[sheet_name]
        ncols = len(df.columns)
        wbord = Side(style="thin", color="FFFFFF")
        border = Border(left=wbord, right=wbord, top=wbord, bottom=wbord)
        top_med = Border(left=wbord, right=wbord, bottom=wbord, top=Side(style="medium", color=HEAD))
        head_fill, data_fill, total_fill = (PatternFill("solid", fgColor=HEAD),
                                            PatternFill("solid", fgColor=DATA),
                                            PatternFill("solid", fgColor=TOTAL))
        white_fill = PatternFill("solid", fgColor=WHITE)
        sold_fill  = PatternFill("solid", fgColor=SOLD)
        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left", vertical="center")

        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            tc = ws.cell(row=1, column=1, value=title)
            tc.fill = head_fill; tc.font = Font(bold=True, color=WHITE, size=13); tc.alignment = left
            ws.row_dimensions[1].height = 24

        hrow = startrow + 1
        for c in ws[hrow]:
            c.fill, c.font, c.alignment, c.border = head_fill, Font(bold=True, color=WHITE), center, border
        ws.freeze_panes = ws.cell(row=hrow + 1, column=1).coordinate
        ws.row_dimensions[hrow].height = 22

        first_data = hrow + 1
        for r in range(first_data, ws.max_row + 1):
            is_total = _is_total_row(ws.cell(row=r, column=1).value)
            di = r - first_data
            is_sold = (sold_mask is not None and not is_total
                       and di < len(sold_mask) and bool(sold_mask[di]))
            for ci in range(1, ncols + 1):
                cell = ws.cell(row=r, column=ci)
                cell.alignment = left if ci == 1 else center
                if ci in aed_idx:
                    cell.number_format = '"AED" #,##0'      # numeric AED currency cells
                if is_total:
                    cell.fill = total_fill
                    cell.font = Font(bold=True, color=DARK)
                    cell.border = top_med
                else:
                    cell.fill = sold_fill if is_sold else (data_fill if di % 2 == 0 else white_fill)
                    cell.font = Font(color=DARK)
                    cell.border = border
        for ci in range(1, ncols + 1):
            col = get_column_letter(ci)
            maxlen = max((len(str(ws.cell(row=rr, column=ci).value or "")) for rr in range(hrow, ws.max_row + 1)),
                         default=12)
            ws.column_dimensions[col].width = min(max(maxlen + 3, 12), 42)
    return out.getvalue()


def export_button(df: pd.DataFrame, file_name, key, title=None, label="⬇️  Export to Excel", **xlsx_kwargs):
    st.download_button(label,
                       data=df_to_styled_xlsx(df, sheet_name=(title or "Sheet1"), title=title, **xlsx_kwargs),
                       file_name=file_name, key=key,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def table_with_export(df: pd.DataFrame, file_name, key, title=None):
    """Render the export button in the top-right above the styled table."""
    c = st.columns([0.74, 0.26])
    with c[1]:
        export_button(df, file_name, key, title)
    excel_table(df)


def aed(x):  return f"AED {x:,.0f}"

def ordinal(n):
    """13 -> '13th', 21 -> '21st'. Accepts ints or numeric strings; passes through non-numeric."""
    try:
        n = int(float(n))
    except (ValueError, TypeError):
        return str(n)
    suf = "th" if 10 <= n % 100 <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"

def area_fmt(x, sqm=False):
    v = x / SQFT_PER_SQM if sqm else x
    return f"{v:,.0f}"


def avail_adjusted_median(sub, col):
    """Median over ALL rows of a typology (Sold + Available), sorted by floor: take the
    median row position (the 52nd of 104, etc.). The value must come from an Available unit —
    if the median row is Sold, step to the row above (lower floor) until an Available one is
    found; if none above, step down. Returns that median row's value in `col` (e.g. PSF_total
    for Median /sqft, or Price for Median Price)."""
    if sub is None or sub.empty:
        return float("nan")
    s = sub.copy()
    s["_fn"] = pd.to_numeric(s["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                             errors="coerce")
    s = s.sort_values(["_fn", "Unit"]).reset_index(drop=True)
    n = len(s)
    mid = (n + 1) // 2 - 1                         # 0-based index of the median row
    if not (s["Status"] == "Available").any():     # fully-sold typology → use the median row as-is
        return float(s.loc[mid, col])
    i = mid
    while i >= 0 and s.loc[i, "Status"] != "Available":
        i -= 1                                      # step up to an Available row
    if i < 0:                                       # none above → step down from the median row
        i = mid
        while i < n and s.loc[i, "Status"] != "Available":
            i += 1
    return float(s.loc[i, col]) if 0 <= i < n else float("nan")


def ensure_new_options(key, options):
    """Keep a 'show-all' multiselect honest: when a brand-new option appears in the data
    (e.g. a freshly added topology), add it to the current selection so it shows by default —
    without resurrecting options the user deliberately unticked."""
    seen_key = f"_seen__{key}"
    seen = st.session_state.setdefault(seen_key, set(options))
    new = [o for o in options if o not in seen]
    seen.update(options)
    if new and key in st.session_state:
        cur = [x for x in st.session_state[key] if x in options]
        st.session_state[key] = cur + [o for o in new if o not in cur]


def column_picker(all_cols, key, locked=None, hidden_default=None):
    """Dropdown (popover) with a multi-select to show / hide table columns.

    `locked` columns are always shown and cannot be unticked.
    `hidden_default` columns start unticked (hidden) but can be shown.
    Returns the ordered list of column names to display.
    """
    all_cols = list(all_cols)
    locked = locked or []
    hidden_default = hidden_default or []
    selectable = [c for c in all_cols if c not in locked]
    default = [c for c in selectable if c not in hidden_default]
    with st.popover("🔧 Columns", use_container_width=False):
        st.caption("Tick to show, untick to hide. Key columns are always shown.")
        chosen = st.multiselect("Columns", selectable, default=default,
                                key=key, label_visibility="collapsed")
    return [c for c in all_cols if c in locked or c in chosen]


# ── Data loading ───────────────────────────────────────────────────────────────

def load_unit_data() -> pd.DataFrame:
    raw  = pd.read_excel(EXCEL_PATH, sheet_name="Muraba Unit Wise Details ", header=None)
    data = raw.iloc[3:].copy().reset_index(drop=True)
    df = pd.DataFrame({
        "Type":          data[0],
        "Status":        data[1],
        "Unit":          data[2].astype(str),
        "Floor":         data[3].astype(str),
        "Parking":       pd.to_numeric(data[4], errors="coerce").fillna(0).astype(int),
        "Internal_sqft": pd.to_numeric(data[5], errors="coerce"),
        "External_sqft": pd.to_numeric(data[6], errors="coerce"),
        "Terrace_Rate":  pd.to_numeric(data[8], errors="coerce"),
        "Price_sqft":    pd.to_numeric(data[10], errors="coerce"),
    })
    df = df[df["Type"].notna() & (df["Type"] != "Total")].reset_index(drop=True)
    df["Type"]   = df["Type"].replace("4 Bedroom Simplex", "4 Bedroom XL")  # renamed typology
    df["Floor"]  = df["Floor"].apply(ordinal)                          # normalise 33 / 33.0 / "4th" -> "33rd" / "4th"
    df["Status"] = df["Status"].replace("Bank Locked", "Available")   # Bank Locked reclassified as Available

    # Fill any missing Price/sqft from the nearest same-type floor (source-data gaps, e.g. unit 6802).
    if df["Price_sqft"].isna().any():
        fn = pd.to_numeric(df["Floor"].str.replace(r"[^0-9]", "", regex=True), errors="coerce")
        for t, idx in df.groupby("Type").groups.items():
            order = fn.loc[idx].sort_values().index            # same-type rows in floor order
            filled = df.loc[order, "Price_sqft"].ffill().bfill()
            df.loc[order, "Price_sqft"] = filled
        df["Price_sqft"] = df["Price_sqft"].fillna(0.0)        # last-resort guard

    df["Terrace_Override"] = pd.NA                                     # per-unit terrace-rate override (set by bulk tool)
    df["Sellable_Override"] = pd.NA                                    # per-unit sellable-area override (set in Edit Units)
    df["Dup_Up"] = pd.NA                                               # loaded duplexes span DOWN; added ones span UP
    df["Adj_Pct"] = pd.NA                                              # per-unit appreciation(+)/discount(-) as a %
    cmts = load_comments_file()                                       # free-text notes per unit, persisted to JSON
    df["Comment"] = [cmts.get(comment_key(u, t, fl), "")
                     for u, t, fl in zip(df["Unit"], df["Type"], df["Floor"])]
    df["uid"] = [f"u{i}" for i in range(len(df))]   # stable unique row id (unit numbers are NOT unique)
    return df


def load_floor_data(units_df: pd.DataFrame) -> list:
    # link Launches-sheet units to register rows by (unit_no, type) — unique in the standard range
    umap = {(str(r["Unit"]), r["Type"]): r["uid"] for _, r in units_df.iterrows()}
    raw = pd.read_excel(EXCEL_PATH, sheet_name="Launches Residences", header=None)
    floors = []
    for i in range(14, 66):
        if i >= len(raw):
            break
        row = raw.iloc[i]
        try:
            fnum = int(float(row[6]))
        except (ValueError, TypeError):
            continue
        if fnum == 0:
            continue
        units = []
        def mk(col_no, col_rate, typ):
            no = str(int(float(row[col_no])))
            return {"unit_no": no, "type": typ, "rate": float(row[col_rate]),
                    "uid": umap.get((no, typ))}
        if pd.notna(row[3]) and pd.notna(row[7]):
            units.append(mk(3, 7, "3 Bedroom"))
        if pd.notna(row[4]) and pd.notna(row[8]):
            units.append(mk(4, 8, "2 Bedroom"))
        if pd.notna(row[5]) and pd.notna(row[9]):
            units.append(mk(5, 9, "2 Bedroom"))
        floors.append({"floor": fnum, "kind": "Standard", "levels": 1, "units": units})
    return floors


def load_pool_floors(units_df: pd.DataFrame) -> list:
    pool = units_df[units_df["Type"].isin(["3 Bedroom Pool", "4 Bedroom Pool"])].copy()
    pool["fn"] = pd.to_numeric(
        pool["Floor"].str.replace("th", "").str.replace("st", "")
                     .str.replace("nd", "").str.replace("rd", "").str.strip(), errors="coerce")
    floors = []
    for fnum, grp in pool.groupby("fn"):
        if pd.isna(fnum):
            continue
        units = [{"unit_no": str(r["Unit"]), "type": r["Type"], "rate": float(r["Price_sqft"]),
                  "uid": r["uid"]} for _, r in grp.iterrows()]
        floors.append({"floor": int(fnum), "kind": "Pool", "levels": 2, "units": units})
    return floors


def build_floor_list(units_df: pd.DataFrame) -> list:
    """Build the floor list directly from the register so EVERY unit is represented
    (Standard, Pool, Duplex, XL, Penthouse). This guarantees the Floor Manager grand
    total matches the portfolio value — earlier it only read the Standard+Pool sheets
    and silently dropped the 4 Duplex/XL/PH units (~AED 825M)."""
    fnum = pd.to_numeric(units_df["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                         errors="coerce")
    buckets = {}
    for pos in range(len(units_df)):
        f = fnum.iloc[pos]
        if pd.isna(f):
            continue
        r = units_df.iloc[pos]
        buckets.setdefault(int(f), []).append({
            "unit_no": str(r["Unit"]), "type": r["Type"],
            "rate": float(r["Price_sqft"]), "uid": r["uid"],
        })
    out = []
    for f in sorted(buckets):
        units = buckets[f]
        types = [u["type"] for u in units]
        levels = max(TYPE_DEFAULTS.get(t, {"levels": 1})["levels"] for t in types)
        if   any("Pool" in t for t in types):    kind = "Pool"
        elif any("Duplex" in t for t in types):  kind = "Duplex"
        elif any(t == "4 Bedroom XL" for t in types): kind = "XL"
        else:                                     kind = "Standard"
        out.append({"floor": f, "kind": kind, "levels": levels, "units": units})
    return out


def load_params() -> dict:
    esc     = dict(ESC_DEFAULTS)
    terrace = dict(TERRACE_DEFAULTS)
    duplex_premium = 0.0
    try:
        lr = pd.read_excel(EXCEL_PATH, sheet_name="Launches Residences", header=None)
        esc["2 Bedroom"]     = float(lr.iloc[3][20])
        esc["3 Bedroom"]     = float(lr.iloc[4][20])
        terrace["standard"]  = float(lr.iloc[6][20])
    except Exception:
        pass
    try:
        xl = pd.read_excel(EXCEL_PATH, sheet_name="Launches   XL   SX   DX", header=None)
        esc["3 Bedroom Pool"]    = float(xl.iloc[1][16])
        esc["4 Bedroom Pool"]    = float(xl.iloc[2][16])
        esc["3 Bedroom Duplex"]  = float(xl.iloc[4][16])
        esc["4 Bedroom XL"] = float(xl.iloc[5][16])
        esc["4 Bedroom Duplex"]  = float(xl.iloc[6][16])
        terrace["3 Bedroom Pool"] = float(xl.iloc[7][16])
        terrace["4 Bedroom Pool"] = float(xl.iloc[8][16])
        terrace["duplex"]         = float(xl.iloc[9][16])
        terrace["simplex"]        = float(xl.iloc[10][16])
        dp = xl.iloc[11][16]
        duplex_premium = float(dp) if pd.notna(dp) else 0.0
    except Exception:
        pass
    esc["3 Bedroom - New"] = esc["2 Bedroom"]   # new type mirrors 2 Bedroom escalation
    area = {t: {"internal": TYPE_DEFAULTS[t]["internal"], "external": TYPE_DEFAULTS[t]["external"]}
            for t in UNIT_TYPES}
    parking = {t: TYPE_DEFAULTS[t]["parking"] for t in UNIT_TYPES}
    return {"escalation": esc, "terrace": terrace, "duplex_premium": duplex_premium, "area": area,
            "parking": parking, "base": {}}


def load_blocked_floors() -> dict:
    """MEP / Majlis levels from the building view — these floor numbers cannot be added."""
    blocked = {}
    try:
        bv = pd.read_excel(EXCEL_PATH, sheet_name="Building  Usman", header=None)
        for _, row in bv.iterrows():
            lvl, desc = row[2], row[3]
            if pd.isna(lvl) or pd.isna(desc):
                continue
            try:
                lvl_int = int(float(lvl))
            except (ValueError, TypeError):
                continue
            d = str(desc).upper()
            if "MEP" in d or "MAJILIS" in d or "MAJLIS" in d:
                blocked[lvl_int] = str(desc).strip()
    except Exception:
        pass
    return blocked


# ── Calculations ───────────────────────────────────────────────────────────────

def terrace_for(t, params):
    tr = params["terrace"]
    if t in ("2 Bedroom", "3 Bedroom - New", "3 Bedroom"): return tr["standard"]
    if t == "3 Bedroom Pool":                           return tr["3 Bedroom Pool"]
    if t == "4 Bedroom Pool":                           return tr["4 Bedroom Pool"]
    if t == "4 Bedroom XL":                        return tr["simplex"]
    if t in ("3 Bedroom Duplex", "4 Bedroom Duplex"):   return tr["duplex"]
    # 5 Bedroom Duplex (penthouse) keeps its own 100% terrace
    return TYPE_DEFAULTS[t]["terrace_rate"]

def escalation_for(t, params):
    return params["escalation"].get(t, 0.0)

def last_available_price(t, units_df):
    """Return the Price_sqft of the highest-floor Available unit for type t (floor-sequence aware).
    Pools price-family types together (3 Bedroom - New rides the 2 Bedroom ladder)."""
    fam = family_types(t)
    sub = units_df[(units_df["Type"].isin(fam)) & (units_df["Status"] == "Available")].copy()
    if sub.empty:
        sub = units_df[units_df["Type"].isin(fam)].copy()
    if sub.empty:
        return 5000.0
    # sort by numeric floor so escalation always references the topmost available unit
    sub["_fnum"] = pd.to_numeric(
        sub["Floor"].str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    sub = sub.sort_values("_fnum")
    return float(sub.iloc[-1]["Price_sqft"])

def _fnum_series(units_df):
    return pd.to_numeric(units_df["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                         errors="coerce")

def escalation_reference(t, target_floor, units_df):
    """Pick the escalation reference for a new/edited unit of type t on target_floor.

    Floor-direction aware (prices always rise with height):

    • If type-t units exist BELOW target → price UP from below:
        A = highest-floor Available unit below; candidates = A + units between A and target;
        reference R = the HIGHEST-priced candidate; steps = distinct floors with a type-t
        unit strictly between R and target, + 1.  rate = R.psf + esc × steps.

    • Else if type-t units exist ABOVE target → price DOWN from above:
        A = lowest-floor Available unit above; candidates = A + units between target and A;
        reference R = the LOWEST-priced candidate; steps = distinct floors with a type-t
        unit strictly between target and R, + 1.  rate = R.psf − esc × steps.

    Returns ({ref_unit, ref_floor, ref_psf, steps}, direction) where direction is +1 / −1,
    or (None, 0) when no comparable type-t unit exists anywhere.
    """
    sub = units_df[units_df["Type"].isin(family_types(t))].copy()   # pool 3BR-New with 2BR ladder
    if sub.empty:
        return None, 0
    sub["fn"] = _fnum_series(sub)
    sub = sub.dropna(subset=["fn"])
    below = sub[sub["fn"] < target_floor]
    above = sub[sub["fn"] > target_floor]

    if not below.empty:                                    # ── price UP from below ──
        avail = below[below["Status"] == "Available"]
        if not avail.empty:
            a_floor = avail["fn"].max()
            cand = below[below["fn"] >= a_floor]
        else:
            cand = below
        R = cand.loc[cand["Price_sqft"].astype(float).idxmax()]
        r_floor = float(R["fn"])
        steps = int(sub[(sub["fn"] > r_floor) & (sub["fn"] < target_floor)]["fn"].nunique()) + 1
        return {"ref_unit": str(R["Unit"]), "ref_floor": int(r_floor),
                "ref_psf": float(R["Price_sqft"]), "steps": steps}, +1

    if not above.empty:                                    # ── price DOWN from above ──
        avail = above[above["Status"] == "Available"]
        if not avail.empty:
            a_floor = avail["fn"].min()
            cand = above[above["fn"] <= a_floor]
        else:
            cand = above
        R = cand.loc[cand["Price_sqft"].astype(float).idxmin()]
        r_floor = float(R["fn"])
        steps = int(sub[(sub["fn"] < r_floor) & (sub["fn"] > target_floor)]["fn"].nunique()) + 1
        return {"ref_unit": str(R["Unit"]), "ref_floor": int(r_floor),
                "ref_psf": float(R["Price_sqft"]), "steps": steps}, -1

    return None, 0

def new_unit_rate(t, target_floor, units_df, params):
    esc = escalation_for(t, params)
    ref, direction = escalation_reference(t, target_floor, units_df)
    if ref is None:
        rate = last_available_price(t, units_df) + esc
    else:
        rate = ref["ref_psf"] + direction * esc * ref["steps"]
    if "Duplex" in t:
        rate += params.get("duplex_premium", 0.0)
    return max(rate, 0.0)

def get_base(t, params):
    """Return the Base Price/sqft set for type t's price-family, or None."""
    b = params.get("base", {})
    for ft in family_types(t):
        if b.get(ft) is not None:
            return float(b[ft])
    return None

def lowest_available_psf(t, units_df):
    """Current lowest-floor Available price/sqft for type t's family."""
    fam = family_types(t)
    sub = units_df[units_df["Type"].isin(fam)].copy()
    if sub.empty:
        return None
    sub["_fn"] = _fnum_series(sub)
    sub = sub.dropna(subset=["_fn"])
    avail = sub[sub["Status"] == "Available"]
    pick = avail if not avail.empty else sub
    return float(pick.sort_values("_fn").iloc[0]["Price_sqft"])

def base_preview(t, base_psf, params):
    """Without applying: how many Available family units would change and the portfolio delta (AED).
    Uses the same anchor as recompute_from_base (lowest Available floor = base)."""
    fam = family_types(t)
    u = st.session_state.units
    fn = _fnum_series(u)
    mask = u["Type"].isin(fam)
    sub = u[mask].assign(_fn=fn[mask]).dropna(subset=["_fn"])
    floors_sorted = sorted(sub["_fn"].unique())
    avail = sub[sub["Status"] == "Available"]
    if avail.empty or not floors_sorted:
        return 0, 0.0
    pos = {f: i for i, f in enumerate(floors_sorted)}
    anchor_pos = pos[avail["_fn"].min()]
    esc = escalation_for(t, params)
    dpx = params.get("duplex_premium", 0.0) if "Duplex" in t else 0.0
    n_change, delta = 0, 0.0
    for idx in sub.index:
        if u.at[idx, "Status"] != "Available":
            continue
        new = max(base_psf + esc * (pos[float(fn.loc[idx])] - anchor_pos) + dpx, 0.0)
        old = float(u.at[idx, "Price_sqft"])
        if abs(new - old) > 1e-9:
            n_change += 1
            tt = u.at[idx, "Type"]
            internal, external = area_for(tt, params)
            delta += (new - old) * (internal + terrace_for(tt, params) * external)
    return n_change, delta

def recompute_from_base(t, base_psf, params):
    """The **lowest Available** unit of the family takes base_psf; every Available unit above it
    follows escalation: rate = base + escalation × (typology floors above the lowest Available
    floor). Sold units stay fixed. Floor steps count existing typology floors, so missing/MEP
    floors don't distort the ladder."""
    fam = family_types(t)
    u = st.session_state.units.copy()
    fn = _fnum_series(u)
    mask = u["Type"].isin(fam)
    sub = u[mask].assign(_fn=fn[mask]).dropna(subset=["_fn"])
    floors_sorted = sorted(sub["_fn"].unique())
    avail = sub[sub["Status"] == "Available"]
    if avail.empty or not floors_sorted:
        return 0
    pos = {f: i for i, f in enumerate(floors_sorted)}
    anchor_pos = pos[avail["_fn"].min()]              # lowest Available floor = base
    esc = escalation_for(t, params)
    dpx = params.get("duplex_premium", 0.0) if "Duplex" in t else 0.0
    changed = 0
    for idx in sub.index:
        if u.at[idx, "Status"] != "Available":
            continue
        steps = pos[float(fn.loc[idx])] - anchor_pos
        rate = max(base_psf + esc * steps + dpx, 0.0)
        if abs(float(u.at[idx, "Price_sqft"]) - rate) > 1e-9:
            u.at[idx, "Price_sqft"] = rate
            changed += 1
    st.session_state.units = u
    return changed

def reladder_typology(t, params):
    """Re-price every **Available** unit of the price-family up the ladder using the current
    escalation (3 Bedroom - New is re-laddered together with 2 Bedroom). If a Base Price is set
    for the family, the ladder is anchored at that base; otherwise anchors are the existing
    entry-price units. Sold units are never touched. Mutates st.session_state.units.
    Returns the number of units whose price changed."""
    base_psf = get_base(t, params)
    if base_psf is not None:                    # base-anchored ladder (per type, lowest floor)
        return recompute_from_base(t, base_psf, params)
    fam = family_types(t)
    u = st.session_state.units.copy()
    fn = _fnum_series(u)
    famask = u["Type"].isin(fam)
    order = (u[famask].assign(_fn=fn[famask])
             .dropna(subset=["_fn"]).sort_values("_fn").index)
    changed = 0
    for idx in order:                       # bottom-up so each unit sees updated floors below it
        if u.at[idx, "Status"] != "Available":
            continue
        tt = u.at[idx, "Type"]
        tf = float(fn.loc[idx])
        ref, direction = escalation_reference(tt, tf, u)
        if ref is None or direction < 0:    # nothing below → this is the anchor, keep its price
            continue
        rate = ref["ref_psf"] + escalation_for(tt, params) * ref["steps"]
        if "Duplex" in tt:
            rate += params.get("duplex_premium", 0.0)
        rate = max(rate, 0.0)
        if abs(float(u.at[idx, "Price_sqft"]) - rate) > 1e-9:
            u.at[idx, "Price_sqft"] = rate
            changed += 1
    st.session_state.units = u
    return changed

def sync_floor_rates():
    """Push current register prices back onto the floor-objects so the Floors table,
    totals and export stay consistent after a reprice."""
    pmap = dict(zip(st.session_state.units["uid"], st.session_state.units["Price_sqft"]))
    for fl in st.session_state.floors:
        for un in fl["units"]:
            if un.get("uid") in pmap:
                un["rate"] = float(pmap[un["uid"]])

def area_for(t, params):
    a = params.get("area", {}).get(t)
    if a:
        return a["internal"], a["external"]
    return TYPE_DEFAULTS[t]["internal"], TYPE_DEFAULTS[t]["external"]

def unit_val(t, rate, params):
    internal, external = area_for(t, params)
    tr = terrace_for(t, params)
    return {"internal": internal*rate, "terrace": external*tr*rate,
            "total": (internal + tr*external)*rate}

def floor_total(fl, params):
    tot = 0.0
    for u in fl["units"]:
        v = unit_val(u["type"], u["rate"], params)["total"]
        tot += 0.0 if pd.isna(v) else v
    return tot

def recalc(df, params):
    df = df.copy()
    pk = params.get("parking", {})
    for t in df["Type"].unique():
        internal, external = area_for(t, params)
        m = df["Type"] == t
        df.loc[m, "Internal_sqft"] = internal
        df.loc[m, "External_sqft"] = external
        df.loc[m, "Terrace_Rate"]  = terrace_for(t, params)
        if t in pk:
            df.loc[m, "Parking"] = int(pk[t])     # configurable parking per typology (cascades)
    # per-unit terrace overrides win over the type default (set by the floor-range tool)
    if "Terrace_Override" in df.columns:
        ov = df["Terrace_Override"].notna()
        df.loc[ov, "Terrace_Rate"] = pd.to_numeric(df.loc[ov, "Terrace_Override"], errors="coerce")
    df["Sellable_sqft"] = df["Internal_sqft"] + df["Terrace_Rate"]*df["External_sqft"]
    df["Total_sqft"]    = df["Internal_sqft"] + df["External_sqft"]
    # per-unit sellable-area override (set in Edit Units) wins over the derived value
    if "Sellable_Override" in df.columns:
        so = df["Sellable_Override"].notna()
        df.loc[so, "Sellable_sqft"] = pd.to_numeric(df.loc[so, "Sellable_Override"], errors="coerce")
    df["Base_Price"]    = df["Price_sqft"] * df["Sellable_sqft"]        # LIST price (pre appreciation/discount)
    # per-unit appreciation(+%) / discount(-%) → the effective price used in the register total,
    # building view and portfolio KPIs. The Base_Price stays the list price for the variance ladder.
    _adj = (pd.to_numeric(df["Adj_Pct"], errors="coerce").fillna(0.0)
            if "Adj_Pct" in df.columns else 0.0)
    df["Price"]         = df["Base_Price"] * (1 + _adj / 100.0)
    return df


# ── Session state ──────────────────────────────────────────────────────────────

def _init():
    _loaded_blocked = None
    if "units" not in st.session_state:
        # last live session state (ephemeral) → committed Base Version (permanent) → original Excel
        loaded = load_state() if has_saved_state() else None
        if loaded is None and has_base():
            loaded = load_base()                     # survives Cloud redeploys/reboots (it's in the repo)
        if loaded is not None:
            units, floors, fparams, ctr, _loaded_blocked = loaded
            st.session_state.units = units
            st.session_state.floors = floors
            st.session_state.fm_params = fparams
            st.session_state.uid_counter = ctr
        else:                                        # fresh from the original Excel
            st.session_state.units = load_unit_data()
            st.session_state.uid_counter = len(st.session_state.units)
        # overlay the committed comments (muraba_comments.json) so saved comments show on EVERY load
        # path (base version / saved state / Excel) — the repo copy is the persistent source of truth
        _cm = load_comments_file()
        if _cm:
            uu = st.session_state.units
            uu["Comment"] = [_cm.get(comment_key(un, ty, fl), cur if isinstance(cur, str) else "")
                             for un, ty, fl, cur in zip(uu["Unit"], uu["Type"], uu["Floor"], uu["Comment"])]
    if "fm_params" not in st.session_state: st.session_state.fm_params = load_params()
    if "floors"    not in st.session_state: st.session_state.floors    = build_floor_list(st.session_state.units)
    if "blocked"   not in st.session_state:
        st.session_state.blocked = _loaded_blocked if _loaded_blocked is not None else load_blocked_floors()

_init()

params  = st.session_state.fm_params
blocked = st.session_state.blocked
df      = recalc(st.session_state.units, params)


# ── Helpers for mutating register + floors ─────────────────────────────────────

def next_uid():
    n = st.session_state.get("uid_counter", 0) + 1
    st.session_state.uid_counter = n
    return f"u{n}"

def gen_unit_nos(floor_num, types_in_order):
    existing = set(st.session_state.units["Unit"].values)
    nos, n = [], 1
    for _ in types_in_order:
        while True:
            cand = str(floor_num*100 + n)
            if cand not in existing and cand not in nos:
                break
            n += 1
        nos.append(cand); n += 1
    return nos

def add_units_to_register(unit_list, floor_num, params):
    """Assigns a fresh uid to each new unit (mutates the dicts) and appends to the register."""
    for u in unit_list:
        uid = next_uid()
        u["uid"] = uid
        d = TYPE_DEFAULTS[u["type"]]
        internal, external = area_for(u["type"], params)
        st.session_state.units = pd.concat([st.session_state.units, pd.DataFrame([{
            "Type": u["type"], "Status": "Available", "Unit": u["unit_no"], "Floor": ordinal(floor_num),
            "Parking": d["parking"], "Internal_sqft": internal, "External_sqft": external,
            "Terrace_Rate": terrace_for(u["type"], params), "Price_sqft": u["rate"],
            "Terrace_Override": pd.NA, "Sellable_Override": pd.NA,
            # a duplex added here is ONE unit that occupies its floor + the floor ABOVE (roof shifts up)
            "Dup_Up": (True if "Duplex" in u["type"] else pd.NA),
            "Adj_Pct": pd.NA,
            "Comment": "", "uid": uid,
        }])], ignore_index=True)

def _digits(s):
    d = "".join(ch for ch in str(s) if ch.isdigit())
    return int(d) if d else None

def _apply_floor_remap(remap, drop_uids=None):
    """Drop the given unit uids, then renumber every remaining floor per `remap` (old→new) across the
    register (Floor + Unit number), the floors list, and the MEP/Majlis map. Keyed by row so it never
    collides. `remap` shifts floors WITHOUT closing empty levels — empty floors are preserved."""
    drop_uids = drop_uids or set()
    if drop_uids:
        st.session_state.units = st.session_state.units[
            ~st.session_state.units["uid"].isin(drop_uids)].reset_index(drop=True)
    u = st.session_state.units
    ufn = pd.to_numeric(u["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    for idx in u.index:
        fv = ufn[idx]
        if pd.isna(fv):
            continue
        of = int(fv)
        if of in remap and remap[of] != of:
            nf = remap[of]; suf = (_digits(u.at[idx, "Unit"]) or 0) % 100
            u.at[idx, "Floor"] = ordinal(nf)
            u.at[idx, "Unit"] = str(nf * 100 + suf)
    for fl in st.session_state.floors:
        of = fl["floor"]
        if of in remap and remap[of] != of:
            nf = remap[of]; fl["floor"] = nf
            for un in fl["units"]:
                suf = (_digits(un.get("unit_no")) or 0) % 100
                un["unit_no"] = str(nf * 100 + suf)
    st.session_state.floors.sort(key=lambda x: x["floor"])
    st.session_state.blocked = {remap.get(k, k): v for k, v in st.session_state.blocked.items()}

def insert_floors_between(N, count, ordered_types, params):
    """Insert `count` new residential floors at N. The residential floors at/above N shift UP by
    `count` non-MEP levels. **All MEP / Majlis floors stay fixed EXCEPT the top one** (default 67),
    which floats so it always sits directly beneath the 5BR penthouse. New floors take the first
    `count` non-MEP levels from N. Returns (new_floor_numbers, remap)."""
    blk = st.session_state.blocked
    top_mep = max(blk) if blk else None                       # the only MEP that floats
    fixed = (set(blk.keys()) - ({top_mep} if top_mep is not None else set())) | {2}
    count = max(1, int(count))
    u = st.session_state.units
    ufn = pd.to_numeric(u["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    res_floors = sorted({int(f) for f in ufn.dropna().astype(int).tolist() if int(f) not in fixed
                         and int(f) != top_mep})
    affected = [f for f in res_floors if f >= N]

    def kth_above(f, k):
        x = f
        while k > 0:
            x += 1
            if x not in fixed:
                k -= 1
        return x

    new_nums, x = [], N
    while len(new_nums) < count:
        while x in fixed:
            x += 1
        new_nums.append(x); x += 1
    remap = {f: kth_above(f, count) for f in affected}        # residential above the cut slides up
    # the top MEP rides up with the penthouse, staying directly beneath the 5BR's two-floor span
    if top_mep is not None and affected:
        pent = max(remap.values())                            # 5BR record after the slide
        new_mep = pent - 2
        new_res = set(remap.values()) | {f for f in res_floors if f not in affected}
        while new_mep in new_res and new_mep > 1:
            new_mep -= 1
        if new_mep != top_mep:
            remap[top_mep] = new_mep
    _apply_floor_remap(remap)

    for nn in new_nums:
        nos = gen_unit_nos(nn, ordered_types)
        new_units = [{"unit_no": no, "type": t,
                      "rate": new_unit_rate(t, nn, st.session_state.units, params)}
                     for no, t in zip(nos, ordered_types)]
        add_units_to_register(new_units, nn, params)
        st.session_state.floors.append({"floor": nn, "kind": "Inserted",
                                        "levels": max(TYPE_DEFAULTS[t]["levels"] for t in ordered_types),
                                        "units": new_units})
    st.session_state.floors.sort(key=lambda x: x["floor"])
    return new_nums, remap

def remove_floors_between(From, To):
    """Remove all residential floors in [From, To]. The residential floors above shift DOWN by
    `count` non-MEP levels. **All MEP / Majlis floors stay fixed EXCEPT the top one** (default 67),
    which floats so it always sits directly beneath the 5BR penthouse. Callers must block this when
    any unit in range is Sold. Returns (removed_floors, remap)."""
    blk = st.session_state.blocked
    top_mep = max(blk) if blk else None                       # the only MEP that floats
    fixed = (set(blk.keys()) - ({top_mep} if top_mep is not None else set())) | {2}
    u = st.session_state.units
    ufn = pd.to_numeric(u["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    res_floors = sorted({int(f) for f in ufn.dropna().astype(int).tolist() if int(f) not in fixed
                         and int(f) != top_mep})
    to_remove = [f for f in res_floors if From <= f <= To]
    above = [f for f in res_floors if f > To]
    if not to_remove:
        return [], {}
    count = len(to_remove)
    rem = set(to_remove)

    def kth_below(f, k):
        x = f
        while k > 0:
            x -= 1
            if x not in fixed:
                k -= 1
        return x

    remap = {f: kth_below(f, count) for f in above}           # residential above the cut slides down
    # the top MEP rides down with the penthouse, staying directly beneath the 5BR's two-floor span
    if top_mep is not None and above:
        pent = max(remap.values())                            # 5BR record after the slide
        new_mep = pent - 2
        new_res = set(remap.values()) | {f for f in res_floors if f not in rem and f not in remap}
        while new_mep in new_res and new_mep > 1:
            new_mep -= 1
        if new_mep != top_mep:
            remap[top_mep] = new_mep
    drop_uids = {u.at[idx, "uid"] for idx in u.index
                 if pd.notna(ufn[idx]) and int(ufn[idx]) in rem}
    _apply_floor_remap(remap, drop_uids)
    return to_remove, remap

def normalize_mep_layout():
    """Rebuild the tower into one clean, contiguous stack from the current register — generic, not
    tied to any particular saved state. The model is a stack of *slabs*:
      • a regular floor          → 1 slab, 1 level
      • a duplex                 → 1 slab, 2 levels (record on the LOWER floor; it owns the empty
                                   floor directly ABOVE and is labelled by its lower floor)
      • each MEP / Majlis floor  → 1 slab, 1 level
      • the amenity floor (2)    → 1 slab, 1 level
    We keep each slab's vertical ORDER, drop every bare gap, re-pack contiguously from the bottom,
    and move the single TOP MEP to sit directly beneath the 5BR penthouse. Fully idempotent: a tower
    that's already clean re-packs to the exact same numbers."""
    u = st.session_state.units
    blk = dict(st.session_state.get("blocked") or {})
    if u.empty:
        return
    if not blk:
        blk = load_blocked_floors()                      # never lose the MEP set
    AMEN = 2

    ufn = pd.to_numeric(u["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    by_floor = {}                                        # floor -> [row indices]
    for idx in u.index:
        if pd.notna(ufn[idx]):
            by_floor.setdefault(int(ufn[idx]), []).append(idx)
    by_floor.pop(AMEN, None)                             # floor 2 is the amenity band

    def _is_dup(idx):
        return "Duplex" in str(u.at[idx, "Type"])

    # Every duplex sits on its (lower) record floor and OWNS the floor directly ABOVE it, which
    # stays empty — the duplex is labelled by its lower floor (e.g. 3BR Duplex 5901 on floor 59
    # occupies 59 + 60). This is universal; the legacy per-unit Dup_Up direction no longer decides
    # the companion.
    dup_floors = {f for f, idxs in by_floor.items() if any(_is_dup(i) for i in idxs)}
    owned = {f + 1: f for f in dup_floors}               # companion (above) -> record floor

    slabs, used = [], set()
    for f in sorted(by_floor):
        if f in used:
            continue
        if f in dup_floors:                              # duplex slab: record floor + empty floor above
            slabs.append({"kind": "dup", "floor": f, "h": 2,
                          "rec": by_floor[f], "comp": by_floor.get(f + 1, [])})
            used.add(f); used.add(f + 1)
        elif f in owned:                                 # a duplex's companion floor (already taken)
            continue
        else:
            slabs.append({"kind": "res", "floor": f, "h": 1, "rec": by_floor[f]})
            used.add(f)
    for f, desc in blk.items():
        slabs.append({"kind": "mep", "floor": f, "h": 1, "desc": desc})
    slabs.append({"kind": "amen", "floor": AMEN, "h": 1})
    slabs.sort(key=lambda s: s["floor"])

    pent = None                                          # the 5BR penthouse slab (highest such)
    for s in slabs:
        if s["kind"] == "dup" and any("5 Bedroom" in str(u.at[i, "Type"]) for i in s["rec"]):
            pent = s
    meps = [s for s in slabs if s["kind"] == "mep"]
    top_mep = meps[-1] if meps else None
    if pent is not None and top_mep is not None:         # top MEP rides directly below the penthouse
        slabs.remove(top_mep)
        slabs.insert(slabs.index(pent), top_mep)

    def _set(idxs, flo):
        for i in idxs:
            suf = (_digits(u.at[i, "Unit"]) or 0) % 100
            u.at[i, "Floor"] = ordinal(flo); u.at[i, "Unit"] = str(flo * 100 + suf)

    new_blk, n = {}, 1
    for s in slabs:
        if s["kind"] == "mep":
            new_blk[n] = s.get("desc", "MEP"); n += 1
        elif s["kind"] == "amen":
            n += 1                                       # reserve floor 2 for amenities (no units)
        elif s["kind"] == "res":
            _set(s["rec"], n); n += 1
        else:                                            # duplex: record on LOWER floor, empty floor ABOVE
            _set(s["rec"], n); _set(s["comp"], n + 1)
            for i in s["rec"]:
                if _is_dup(i) and "Dup_Up" in u.columns:
                    u.at[i, "Dup_Up"] = True             # owns the floor above (labelled by lower floor)
            n += 2
    st.session_state.blocked = new_blk
    st.session_state.floors = build_floor_list(st.session_state.units)

def remove_units_from_register(uids):
    st.session_state.units = st.session_state.units[
        ~st.session_state.units["uid"].isin(uids)].reset_index(drop=True)

def mix_from_editor(edited):
    out = []
    for r in edited.itertuples():
        if pd.notna(r.Type) and pd.notna(r.Qty) and int(r.Qty) > 0:
            out.append((str(r.Type), int(r.Qty)))
    return out

def uid_status_map():
    u = st.session_state.units
    return dict(zip(u["uid"], u["Status"]))

def unit_status(u, smap):
    return smap.get(u.get("uid"), "Available")

def split_floor_units(fl):
    """Split a floor's units into locked (Sold/Bank Locked) and editable (Available), by uid."""
    smap = uid_status_map()
    locked, avail = [], []
    for u in fl["units"]:
        (avail if unit_status(u, smap) == "Available" else locked).append(u)
    return locked, avail

def clear_builder(state_key):
    for k in list(st.session_state.keys()):
        if k.startswith(state_key):
            st.session_state.pop(k, None)

def unit_mix_builder(state_key, default_rows, qmin=1):
    """Rows of (Type dropdown + Qty stepper). Returns list of (type, qty). Uses +/- steppers."""
    sk_rows, sk_ctr = f"{state_key}__rows", f"{state_key}__ctr"
    if sk_rows not in st.session_state:
        st.session_state[sk_ctr] = 0
        st.session_state[sk_rows] = []
        for r in default_rows:
            st.session_state[sk_rows].append({"id": st.session_state[sk_ctr], "type": r["type"], "qty": r["qty"]})
            st.session_state[sk_ctr] += 1
    rows = st.session_state[sk_rows]
    h1, h2, h3 = st.columns([3, 2, 1])
    h1.caption("Topology"); h2.caption("Quantity (use − / +)"); h3.caption(" ")
    to_del = None
    for r in rows:
        rid = r["id"]
        c1, c2, c3 = st.columns([3, 2, 1])
        r["type"] = c1.selectbox("Type", UNIT_TYPES,
                                 index=UNIT_TYPES.index(r["type"]) if r["type"] in UNIT_TYPES else 0,
                                 key=f"{state_key}__t{rid}", label_visibility="collapsed")
        r["qty"] = c2.number_input("Qty", min_value=qmin, max_value=50, value=max(int(r["qty"]), qmin),
                                   step=1, key=f"{state_key}__q{rid}", label_visibility="collapsed")
        if c3.button("🗑️", key=f"{state_key}__x{rid}", help="Remove this row"):
            to_del = rid
    if to_del is not None:
        st.session_state[sk_rows] = [r for r in rows if r["id"] != to_del]
        st.session_state.pop(f"{state_key}__t{to_del}", None)
        st.session_state.pop(f"{state_key}__q{to_del}", None)
        st.rerun()
    if st.button("➕ Add topology", key=f"{state_key}__add"):
        st.session_state[sk_rows].append({"id": st.session_state[sk_ctr], "type": "2 Bedroom", "qty": qmin})
        st.session_state[sk_ctr] += 1
        st.rerun()
    return [(r["type"], int(r["qty"])) for r in st.session_state[sk_rows]]


# One-time on load: keep MEP floors visible (clear any MEP buried under a duplex; park the top MEP
# directly below the 5BR penthouse). Idempotent — a clean tower is untouched.
if not st.session_state.get("_mep_normalized"):
    normalize_mep_layout()
    st.session_state["_mep_normalized"] = True
    blocked = st.session_state.blocked
    df = recalc(st.session_state.units, params)


# ── Sidebar ────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.title("Muraba Veil")
    st.caption("Unit Manager")

    st.caption("✅ Every edit applies **instantly** across all tabs. To keep a version, use "
               "**Save Base Version** below — it's stored in the cloud and persists across restarts.")

    if st.button("↩️  Reset to original Excel", use_container_width=True):
        clear_saved_state()
        for k in ["units", "fm_params", "floors", "blocked", "uid_counter"]:
            st.session_state.pop(k, None)
        st.session_state["flash"] = ("success", "↩️ Reset to the original Excel baseline.")
        st.rerun()

    st.divider()
    st.caption("**Base Versions** — save the current state under a name, or load any saved version. "
               "Stored in the cloud database (Supabase), so they persist across restarts and "
               "redeploys. Each Save / Load asks for the app password.")
    _app_pwd = st.secrets.get("password", os.environ.get("APP_PASSWORD", "muraba2026"))

    if not sb_enabled():
        st.info("Base versions need Supabase configured — add `supabase_url` and `supabase_key` "
                "to the app secrets.")
    else:
        # ── Save Base Version → app password, then a name ──
        if st.button("📌  Save Base Version", use_container_width=True, key="bv_save_btn"):
            st.session_state["bv_save_prompt"] = True
            st.session_state["bv_load_prompt"] = False
            st.session_state.pop("bv_save_pwd", None)
            st.session_state.pop("bv_save_name", None)
        if st.session_state.get("bv_save_prompt"):
            _sp = st.text_input("App password", type="password", key="bv_save_pwd")
            _sname = st.text_input("Name this base version", key="bv_save_name",
                                   placeholder="e.g. Pre-launch · Mar 2026")
            st.caption("Saving under an existing name overwrites that version.")
            _sc1, _sc2 = st.columns(2)
            if _sc1.button("Confirm Save", use_container_width=True, key="bv_save_ok",
                           disabled=not _sname.strip()):
                if _sp != _app_pwd:
                    st.error("Incorrect password.")
                else:
                    try:
                        sb_save_version(_sname.strip(), _state_dict())
                        st.session_state["nv_nonce"] = st.session_state.get("nv_nonce", 0) + 1
                        st.session_state["bv_save_prompt"] = False
                        st.session_state.pop("bv_save_pwd", None)
                        st.session_state.pop("bv_save_name", None)
                        st.session_state["flash"] = ("success", f"📌 Saved base version “{_sname.strip()}”.")
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Save failed: {_e}")
            if _sc2.button("Cancel", use_container_width=True, key="bv_save_cancel"):
                st.session_state["bv_save_prompt"] = False
                st.session_state.pop("bv_save_pwd", None)
                st.session_state.pop("bv_save_name", None)
                st.rerun()

        # ── Load Base Version → app password, then pick from the list, then confirm ──
        if st.button("📥  Load Base Version", use_container_width=True, key="bv_load_btn"):
            st.session_state["bv_load_prompt"] = True
            st.session_state["bv_save_prompt"] = False
            st.session_state.pop("bv_load_pwd", None)
        if st.session_state.get("bv_load_prompt"):
            _lp = st.text_input("App password", type="password", key="bv_load_pwd")
            _names = [n for n, _ in sb_list_versions(st.session_state.get("nv_nonce", 0))]
            _sel = st.selectbox("Choose a base version to load", _names, key="bv_load_select") if _names else None
            if not _names:
                st.caption("No base versions saved yet.")
            _lc1, _lc2 = st.columns(2)
            if _lc1.button("✅  Confirm Load", use_container_width=True, key="bv_load_ok",
                           disabled=not _names):
                if _lp != _app_pwd:
                    st.error("Incorrect password.")
                else:
                    try:
                        _state = sb_load_version(_sel)
                    except Exception as _e:
                        _state = None
                        st.error(f"Load failed: {_e}")
                    if _state:
                        _u, _fl, _fp, _ctr, _blk = _parse_state(_state)
                        st.session_state.units = _u
                        st.session_state.floors = _fl
                        st.session_state.fm_params = _fp
                        st.session_state.uid_counter = _ctr
                        st.session_state.blocked = _blk if _blk else load_blocked_floors()
                        normalize_mep_layout()
                        st.session_state["bv_load_prompt"] = False
                        st.session_state.pop("bv_load_pwd", None)
                        st.session_state["flash"] = ("success", f"📥 Loaded base version “{_sel}”.")
                        st.rerun()
                    elif _lp == _app_pwd:
                        st.error("That version could not be found.")
            if _lc2.button("Cancel", use_container_width=True, key="bv_load_cancel"):
                st.session_state["bv_load_prompt"] = False
                st.session_state.pop("bv_load_pwd", None)
                st.rerun()
            if _sel:
                st.caption(f"Will load: **{_sel}**")

    st.divider()
    st.caption("Add / edit / remove floors in the **Floor Manager** tab. Changes show everywhere "
               "immediately; use **Save Base Version** to persist them across restarts.")
    if blocked:
        st.caption("**Blocked floors (MEP / Majlis):** " + ", ".join(str(k) for k in sorted(blocked)))


# ── Top KPIs (full values) ─────────────────────────────────────────────────────

st.markdown("""<style>
[data-testid="stMetricValue"] { font-size: 1.35rem; }
[data-testid="stMetricLabel"] { font-size: 0.85rem; }
</style>""", unsafe_allow_html=True)

st.title("Muraba Veil — Unit Manager")

_flash = st.session_state.pop("flash", None)
if _flash:
    getattr(st, _flash[0])(_flash[1])

# Let metric labels & values wrap / scale instead of truncating with an ellipsis on
# narrow windows (otherwise long values like "AED 5,161,…" get cut off).
st.markdown(
    """<style>
    [data-testid="stMetricValue"], [data-testid="stMetricValue"] > div,
    [data-testid="stMetricLabel"], [data-testid="stMetricLabel"] p,
    [data-testid="stMetricLabel"] > div{
        white-space:normal; overflow:visible; text-overflow:clip;}
    [data-testid="stMetricValue"]{font-size:clamp(1.05rem,2.1vw,1.9rem); line-height:1.2;}
    </style>""",
    unsafe_allow_html=True,
)

# Global scorecards (full building) — shown on top of every page, no per-tab duplication
ALLOWABLE_SELLABLE = 818186.683338944          # fixed design cap; shown rounded as 818,187
_tot_area = df["Total_sqft"].sum()
_variance = _tot_area - ALLOWABLE_SELLABLE
g1, g2, g3, g4, g5 = st.columns(5)
g1.metric("Units shown", len(df))
g2.metric("Total Area (sqft)", f"{_tot_area:,.0f}")
g3.metric("Total Allowable Sellable (sqft)", f"{ALLOWABLE_SELLABLE:,.0f}")
g4.metric("Variance: Total − Allowable (sqft)", f"{_variance:,.0f}",
          delta=f"{_variance:,.0f}", delta_color="inverse")
g5.metric("Total Price/sqft", aed(df["Price"].sum()/_tot_area) if _tot_area else "—")
# Portfolio Value on its own line below the rest (wider column so the full AED value fits)
gp = st.columns([2, 1, 1, 1])
gp[0].metric("Portfolio Value", aed(df["Price"].sum()))

st.divider()

# Both Building View tabs are shown by default (for stakeholders to compare/decide).
# To hide them later, set `show_building_view = false` in secrets or SHOW_BUILDING_VIEW=0.
try:
    _bv_secret = st.secrets.get("show_building_view", "")
except Exception:
    _bv_secret = ""
_bv_flag = str(_bv_secret or os.environ.get("SHOW_BUILDING_VIEW", "")).strip().lower()
SHOW_BV = _bv_flag not in ("0", "false", "no", "off")

_labels = ["Unit Register", "Summary by Type", "Topology View"]
if SHOW_BV:
    _labels += ["Muraba Veil - Building View"]
_labels += ["Floor Manager", "Edit / Remove Units"]
_tmap = dict(zip(_labels, st.tabs(_labels)))
tab1 = _tmap["Unit Register"]; tab2 = _tmap["Summary by Type"]; tab5 = _tmap["Topology View"]
tab3 = _tmap["Floor Manager"]; tab4 = _tmap["Edit / Remove Units"]
tab6 = None                          # legacy dark Building View — hidden
tab6b = None                         # legacy enhanced (✦) Building View — hidden
tab6c = _tmap.get("Muraba Veil - Building View")


# ── Tab 1: Unit Register ───────────────────────────────────────────────────────

with tab1:
    # Per-floor LIST-price step vs the SAME COLUMN one floor below in the same typology.
    # Match by (Type, unit-number suffix) so 1703 compares to 1603 (not its same-floor twin), use
    # AVAILABLE units only (so a sold/discounted neighbour can't distort it), and divide by the floor
    # gap when the immediate floor below is sold — giving the true per-floor escalation either way.
    order = df.copy()
    order["fnum"] = pd.to_numeric(order["Floor"].str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    order["suf"]  = pd.to_numeric(order["Unit"].str.replace(r"[^0-9]", "", regex=True), errors="coerce") % 100
    # Build each column (same Type + unit-number suffix) as a STACK of its units ordered by floor.
    # Consecutive units in a stack are ONE price step apart no matter how many MEP/empty floors lie
    # between them in number (those carry no units) — so we step by STACK POSITION, never by raw floor
    # numbers, and skip Sold neighbours (no list price), dividing by the positions skipped.
    _stacks = {}
    for _, _r in order.iterrows():
        if pd.isna(_r["fnum"]) or pd.isna(_r["suf"]):
            continue
        _stacks.setdefault((_r["Type"], int(_r["suf"])), []).append(
            (int(_r["fnum"]), _r["uid"], float(_r["Base_Price"]), float(_r["Price_sqft"]),
             _r["Status"] == "Available"))
    esc_map, var_map = {}, {}
    for _lst in _stacks.values():
        _lst.sort()
        _avail = [k for k, _t in enumerate(_lst) if _t[4]]   # stack positions of Available units
        for _i, (_f, _uid, _price, _psf, _av) in enumerate(_lst):
            if not _av:                                   # Sold (or N/A) → blank
                esc_map[_uid] = var_map[_uid] = pd.NA
                continue
            _below = [k for k in _avail if k < _i]
            _above = [k for k in _avail if k > _i]
            if _below:                                    # prefer the nearest Available unit BELOW
                _j = max(_below); _steps = _i - _j
                esc_map[_uid] = (_psf - _lst[_j][3]) / _steps
                var_map[_uid] = (_price - _lst[_j][2]) / _steps
            elif _above:                                  # else fall back to the nearest one ABOVE (same step)
                _j = min(_above); _steps = _j - _i
                esc_map[_uid] = (_lst[_j][3] - _psf) / _steps
                var_map[_uid] = (_lst[_j][2] - _price) / _steps
            else:                                         # only Available unit in its column → blank
                esc_map[_uid] = var_map[_uid] = pd.NA

    type_opts = [t for t in UNIT_TYPES if t in set(df["Type"])] + \
                [t for t in sorted(df["Type"].unique()) if t not in UNIT_TYPES]
    ensure_new_options("reg_type_filter", type_opts)
    fc1, fc2 = st.columns(2)
    f_types  = fc1.multiselect("Type", type_opts, default=type_opts, key="reg_type_filter")
    f_status = fc2.multiselect("Status", STATUS_OPTIONS, default=STATUS_OPTIONS, key="reg_status_filter")
    view = df[df["Type"].isin(f_types) & df["Status"].isin(f_status)].copy()

    # Default sort: by typology (topology order), then by unit number (…02 before …03, etc.)
    _trank = {t: i for i, t in enumerate(UNIT_TYPES)}
    view["_tr"]   = view["Type"].map(_trank).fillna(999)
    view["_unum"] = pd.to_numeric(view["Unit"].str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    view = view.sort_values(["_tr", "_unum", "Unit"]).drop(columns=["_tr", "_unum"])

    # Derived per-unit columns
    view["PSF_total"]  = view["Price"] / view["Total_sqft"]
    view["Int_Value"]  = view["Price_sqft"] * view["Internal_sqft"]
    view["Terr_Value"] = view["Price_sqft"] * view["Terrace_Rate"] * view["External_sqft"]
    view["Esc_row"]    = view["uid"].map(esc_map)
    view["Var_row"]    = view["uid"].map(var_map)
    view.loc[view["Status"] == "Sold", "Var_row"] = pd.NA   # Floor Wise Variance blank for Sold units

    if "Comment" not in view.columns:
        view["Comment"] = ""

    def _money(v):  return "" if pd.isna(v) else f"AED {v:,.0f}"
    def _num1(v):   return "" if pd.isna(v) else f"{v:,.1f}"
    def _adjpct(v):                                    # signed % for appreciation(+)/discount(-); blank if none
        v = pd.to_numeric(v, errors="coerce")
        if pd.isna(v) or round(float(v), 4) == 0:
            return ""
        return ("+" if v > 0 else "−") + f"{abs(float(v)):.1f}%"

    # Pre-format every value to a string so the look matches; Comment stays editable.
    disp = pd.DataFrame({
        "Type": view["Type"].values, "Status": view["Status"].values,
        "Unit": view["Unit"].values, "Floor": view["Floor"].values,
        "Parking": view["Parking"].astype(int).astype(str).values,
        "Internal (sqft)": view["Internal_sqft"].map(_num1).values,
        "External (sqft)": view["External_sqft"].map(_num1).values,
        "Total Area (sqft)": view["Total_sqft"].map(_num1).values,
        "Sellable (sqft)": view["Sellable_sqft"].map(_num1).values,
        "Terrace Rate": view["Terrace_Rate"].map(lambda v: "" if pd.isna(v) else f"{v:.0%}").values,
        "Price/Sellable sqft": view["Price_sqft"].map(_money).values,
        "Price/Total sqft": view["PSF_total"].map(_money).values,
        "Internal Value (AED)": view["Int_Value"].map(_money).values,
        "Terrace Value (AED)": view["Terr_Value"].map(_money).values,
        "List Price (AED)": view["Base_Price"].map(_money).values,
        "Appreciation / Discount": view["Adj_Pct"].map(_adjpct).values,
        "Final Price (AED)": view["Price"].map(_money).values,
        "Escalation vs below (/sqft)": view["Esc_row"].map(_money).values,
        "Floor Wise Variance (AED)": view["Var_row"].map(_money).values,
        "Comment": view["Comment"].fillna("").astype(str).values,
        "uid": view["uid"].values,
    })
    display_cols = [c for c in disp.columns if c != "uid"]
    show_cols = column_picker(display_cols, key="reg_cols", locked=["Type", "Unit"])

    # Styled read-only table with the blue Sold highlight (reliable, no extra deps)
    sold_by_idx = disp.set_index("uid")["Status"] == "Sold"
    vis = disp[show_cols].copy()
    vis.index = disp["uid"].values
    def _hl_sold(row):
        return ["background-color:#9DC3E6" if bool(sold_by_idx.loc[row.name]) else "" for _ in row]
    def _color_adj(val):                               # green = appreciation, red = discount
        s = str(val)
        if s.startswith("+"):                          return "color:#1a7f37;font-weight:600"
        if s.startswith("−") or s.startswith("-"):     return "color:#d1242f;font-weight:600"
        return ""
    _ur = st.columns([0.74, 0.26])
    with _ur[1]:
        # export AED columns as real numbers (currency format) + paint Sold rows yellow
        _reg_raw = {
            "Price/Sellable sqft": view["Price_sqft"], "Price/Total sqft": view["PSF_total"],
            "Internal Value (AED)": view["Int_Value"], "Terrace Value (AED)": view["Terr_Value"],
            "List Price (AED)": view["Base_Price"], "Final Price (AED)": view["Price"],
            "Escalation vs below (/sqft)": view["Esc_row"],
            "Floor Wise Variance (AED)": view["Var_row"],
        }
        _exp = vis.reset_index(drop=True).copy()
        _aed_present = [c for c in _exp.columns if c in _reg_raw]
        for c in _aed_present:
            _exp[c] = pd.to_numeric(pd.Series(_reg_raw[c]).values, errors="coerce")
        export_button(_exp, "Unit_Register.xlsx", key="exp_reg",
                      title="Muraba Veil Unit Register",
                      aed_cols=_aed_present, sold_mask=(view["Status"] == "Sold").values)
    # size the Comment column to its longest text so full comments are readable (the table scrolls
    # horizontally for the overflow); other columns keep their natural width
    _colcfg = None
    if "Comment" in vis.columns and len(vis):
        _ml = int(vis["Comment"].astype(str).map(len).max())
        _colcfg = {"Comment": st.column_config.TextColumn("Comment", width=max(180, min(720, _ml * 7 + 24)))}
    _sty = vis.style.apply(_hl_sold, axis=1)
    if "Appreciation / Discount" in vis.columns:
        _sty = _sty.map(_color_adj, subset=["Appreciation / Discount"])
    st.dataframe(_sty, use_container_width=True,
                 hide_index=True, height=460, column_config=_colcfg)
    st.caption(f"Showing {len(view)} of {len(df)} units · Sold units highlighted in blue · "
               f"Escalation / Floor-Wise Variance = the per-floor list-price step vs the same column "
               f"one floor below (Available units only; blank for Sold)")

    # Inline comment editor (toggle to show/hide); edits persist to file and reload on launch
    show_editor = st.toggle("✏️ Show inline comment editor", value=False, key="show_cmt_editor")
    if show_editor:
        cdf = disp[["Unit", "Type", "Floor", "Status", "Comment", "uid"]].copy()
        ed = st.data_editor(
            cdf, hide_index=True, use_container_width=True, key="cmt_editor",
            column_order=["Unit", "Type", "Floor", "Status", "Comment"],
            disabled=["Unit", "Type", "Floor", "Status"],
            column_config={"Comment": st.column_config.TextColumn("Comment", width="large")},
        )
        # export comments only (keyed by unit|type|floor) — for preserving them across redeploys
        _cmts = {}
        for _, _r in ed.iterrows():
            _c = "" if pd.isna(_r["Comment"]) else str(_r["Comment"]).strip()
            if _c:
                _cmts[comment_key(_r["Unit"], _r["Type"], _r["Floor"])] = _c
        _cd1, _cd2 = st.columns([0.32, 0.68])
        _cd1.download_button("⬇️  Download comments (.json)", disabled=not _cmts,
                             data=json.dumps(_cmts, ensure_ascii=False, indent=2).encode("utf-8"),
                             file_name="muraba_comments.json", mime="application/json",
                             use_container_width=True, key="dl_comments")
        _cd2.caption(f"{len(_cmts)} comment(s) — send this file to persist them across redeploys.")
        new_cmt = ed["Comment"].fillna("").astype(str).values
        if not (new_cmt == cdf["Comment"].values).all():
            cmap = dict(zip(ed["uid"].values, new_cmt))
            u = st.session_state.units
            u["Comment"] = u.apply(lambda r: cmap.get(r["uid"], r.get("Comment", "")), axis=1)
            persist_all_comments()
            st.rerun()


# ── Tab 2: Summary by Type (no Bank Locked column, full values) ────────────────

with tab2:
    st.caption("Mirrors the Excel **Muraba Veil Sale Summary** tab (same columns & structure). "
               "Based on **all units** (Sold included) and computed live from the unit register "
               "using the app's sellable-area pricing.")

    # Column order exactly as in the Excel Sale Summary tab
    SUM_COLS = ["Typology", "Number of Units", "Avg. Price /Sq.ft", "Area (sqft)", "Area (sqm)",
                "Internal (sqft/unit)", "Terrace (sqft/unit)", "Total Internal (sqft)",
                "Total Terrace (sqft)", "Total Sellable", "Counted Terraces",
                "Total Sellable Counted", "Avg Price (per unit)", "Total Sales",
                "Parking", "Total Parking"]

    # build one numeric row per typology, ordered like the master type list
    present = [t for t in UNIT_TYPES if t in set(df["Type"])]
    present += [t for t in sorted(df["Type"].unique()) if t not in present]
    num_rows = []
    for t in present:
        g = df[df["Type"] == t]
        n = len(g)
        internal_u = g["Internal_sqft"].mean()                 # constant per type after recalc
        terrace_u  = g["External_sqft"].mean()
        area_sqft  = internal_u + terrace_u                     # full footprint per unit
        tot_internal = g["Internal_sqft"].sum()
        tot_terrace  = g["External_sqft"].sum()                 # 100% terrace footprint
        tot_sellable = g["Total_sqft"].sum()                   # Internal + full Terrace
        tot_counted  = g["Sellable_sqft"].sum()                # rate-adjusted (app) sellable
        counted_terr = tot_counted - tot_internal              # terrace portion actually counted
        total_sales  = g["Price"].sum()
        num_rows.append({
            "Typology": t,
            "Number of Units": n,
            "Avg. Price /Sq.ft": total_sales / tot_sellable if tot_sellable else 0.0,
            "Area (sqft)": area_sqft,
            "Area (sqm)": area_sqft / SQFT_PER_SQM,
            "Internal (sqft/unit)": internal_u,
            "Terrace (sqft/unit)": terrace_u,
            "Total Internal (sqft)": tot_internal,
            "Total Terrace (sqft)": tot_terrace,
            "Total Sellable": tot_sellable,
            "Counted Terraces": counted_terr,
            "Total Sellable Counted": tot_counted,
            "Avg Price (per unit)": total_sales / n if n else 0.0,
            "Total Sales": total_sales,
            "Parking": int(g["Parking"].mode().iloc[0]) if not g["Parking"].mode().empty else 0,
            "Total Parking": int(g["Parking"].sum()),
        })
    nm = pd.DataFrame(num_rows)

    # Total row (sum additive; Avg. Price /Sq.ft value-weighted; per-unit cells blank like Excel)
    tot_counted_all = nm["Total Sellable Counted"].sum()
    tot_sellable_all = nm["Total Sellable"].sum()
    total = {
        "Typology": "Total",
        "Number of Units": int(nm["Number of Units"].sum()),
        "Avg. Price /Sq.ft": (nm["Total Sales"].sum() / tot_sellable_all) if tot_sellable_all else 0.0,
        "Area (sqft)": None, "Area (sqm)": None,
        "Internal (sqft/unit)": None, "Terrace (sqft/unit)": None,
        "Total Internal (sqft)": nm["Total Internal (sqft)"].sum(),
        "Total Terrace (sqft)": nm["Total Terrace (sqft)"].sum(),
        "Total Sellable": nm["Total Sellable"].sum(),
        "Counted Terraces": nm["Counted Terraces"].sum(),
        "Total Sellable Counted": tot_counted_all,
        "Avg Price (per unit)": None,
        "Total Sales": nm["Total Sales"].sum(),
        "Parking": None,
        "Total Parking": int(nm["Total Parking"].sum()),
    }
    nm = pd.concat([nm, pd.DataFrame([total])], ignore_index=True)

    # formatting per column
    def _f(v, kind):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        if kind == "int":   return f"{int(v):,}"
        if kind == "area":  return f"{v:,.2f}"
        if kind == "areak": return f"{v:,.0f}"
        if kind == "aed0":  return f"AED {v:,.0f}"
        return str(v)

    disp = pd.DataFrame({
        "Typology": nm["Typology"],
        "Number of Units": nm["Number of Units"].apply(lambda v: _f(v, "int")),
        "Avg. Price /Sq.ft": nm["Avg. Price /Sq.ft"].apply(lambda v: _f(v, "aed0")),
        "Area (sqft)": nm["Area (sqft)"].apply(lambda v: _f(v, "area")),
        "Area (sqm)": nm["Area (sqm)"].apply(lambda v: _f(v, "area")),
        "Internal (sqft/unit)": nm["Internal (sqft/unit)"].apply(lambda v: _f(v, "area")),
        "Terrace (sqft/unit)": nm["Terrace (sqft/unit)"].apply(lambda v: _f(v, "area")),
        "Total Internal (sqft)": nm["Total Internal (sqft)"].apply(lambda v: _f(v, "areak")),
        "Total Terrace (sqft)": nm["Total Terrace (sqft)"].apply(lambda v: _f(v, "areak")),
        "Total Sellable": nm["Total Sellable"].apply(lambda v: _f(v, "areak")),
        "Counted Terraces": nm["Counted Terraces"].apply(lambda v: _f(v, "areak")),
        "Total Sellable Counted": nm["Total Sellable Counted"].apply(lambda v: _f(v, "areak")),
        "Avg Price (per unit)": nm["Avg Price (per unit)"].apply(lambda v: _f(v, "aed0")),
        "Total Sales": nm["Total Sales"].apply(lambda v: _f(v, "aed0")),
        "Parking": nm["Parking"].apply(lambda v: _f(v, "int")),
        "Total Parking": nm["Total Parking"].apply(lambda v: _f(v, "int")),
    })[SUM_COLS]

    sum_show = column_picker(list(disp.columns), key="sum_cols", locked=["Typology"])
    table_with_export(disp[sum_show], "Muraba_Veil_Sale_Summary.xlsx", "exp_sum",
                      title="Muraba Veil Sale Summary")
    st.caption(f"Conversion: 1 m² = {SQFT_PER_SQM} ft²  ·  "
               "Total Sellable = Internal + full Terrace  ·  "
               "Counted Terraces = rate-adjusted terrace  ·  "
               "Total Sellable Counted = Internal + Counted Terraces (drives Price per unit).")

    st.divider()

    # ── Furniture Pack (static reference, from Excel Sale Summary) ──────────────
    with st.expander("🛋️  Muraba Veil Furniture Pack (reference, from Excel)", expanded=False):
        fp = pd.DataFrame({
            "Type": ["2 Bedroom", "3 Bedroom", "3 Bedroom Pool", "3 Bedroom XL",
                     "3 Bedroom Duplex", "4 Bedroom", "4 Bedroom Duplex", "5 Bedroom PH"],
            "Amount in AED": [475000, 550000, 600000, 800000, 800000, 850000, 850000, 1250000],
        })
        fp["Amount in AED"] = fp["Amount in AED"].apply(lambda x: f"AED {x:,.0f}")
        table_with_export(fp, "Furniture_Pack.xlsx", "exp_fp", title="Muraba Veil Furniture Pack")


# ── Tab 5: Topology View (min/max/avg stats) ───────────────────────────────────

with tab5:
    st.subheader("Topology Summary Statistics")
    st.caption("**Total Units**, **Total Value** and **Avg Unit Price** (= Total Value ÷ Total Units) "
               "include all units (Sold + Available). **Median /sqft** takes the median row position "
               "across all rows (incl. Sold) but reports the nearest **Available** unit's value "
               "(stepping up if the median row is Sold). **Min / Max / Avg /sqft** are Available-only — "
               "except a typology with **no Available units**, which falls back to its Sold units so it "
               "still appears. Price/sqft = unit Price ÷ Total Area (Internal + External).")
    all_types = [t for t in UNIT_TYPES if t in set(df["Type"])] + \
                [t for t in sorted(df["Type"].unique()) if t not in UNIT_TYPES]
    ensure_new_options("topo_filter", all_types)
    pick = st.multiselect("Filter topologies", all_types, default=all_types, key="topo_filter")

    # all-status aggregate (Sold + Available) → drives Total Units, Total Value, Avg, Median
    alldf = df.copy()
    if pick:
        alldf = alldf[alldf["Type"].isin(pick)]
    alldf["PSF_total"] = alldf["Price"] / alldf["Total_sqft"]
    median_psf   = {t: avail_adjusted_median(g, "PSF_total") for t, g in alldf.groupby("Type")}
    median_price = {t: avail_adjusted_median(g, "Price")     for t, g in alldf.groupby("Type")}

    if alldf.empty:
        st.info("No units for the selected topologies.")
    else:
        # Per-type stats. Min/Max/Avg-per-sqft use Available units; a FULLY-SOLD typology
        # falls back to its Sold units so it still shows (other typologies are unchanged).
        stat_rows = []
        for t, g_all in alldf.groupby("Type"):
            avail = g_all[g_all["Status"] == "Available"]
            base = avail if not avail.empty else g_all      # fully-sold → use sold units
            base_area = base["Total_sqft"].sum()
            n_all = len(g_all)
            stat_rows.append({
                "Type": t,
                "Total_Units": n_all,
                "Min_PSF": base["PSF_total"].min(),
                "Max_PSF": base["PSF_total"].max(),
                "Avg_PSF": (base["Price"].sum() / base_area) if base_area else float("nan"),
                "Median_PSF": median_psf.get(t, float("nan")),
                "Median_Price": median_price.get(t, float("nan")),
                "Min_Price": base["Price"].min(),
                "Max_Price": base["Price"].max(),
                "Total_Value_All": g_all["Price"].sum(),
                "Avg_Price": (g_all["Price"].sum() / n_all) if n_all else float("nan"),
            })
        tv = pd.DataFrame(stat_rows)
        # preserve topology order
        _o = {t: i for i, t in enumerate(all_types)}
        tv = tv.sort_values("Type", key=lambda s: s.map(_o)).reset_index(drop=True)

        tvd = tv[["Type","Total_Units","Min_PSF","Median_PSF","Max_PSF",
                  "Min_Price","Median_Price","Max_Price","Avg_Price","Total_Value_All"]].copy()
        for c in ["Min_PSF","Median_PSF","Max_PSF"]:
            tvd[c] = tvd[c].apply(lambda x: f"AED {x:,.0f}")
        for c in ["Min_Price","Median_Price","Max_Price","Avg_Price","Total_Value_All"]:
            tvd[c] = tvd[c].apply(lambda x: aed(x))
        tvd["Total_Units"]     = tvd["Total_Units"].astype(int)
        tvd.columns = ["Type","Total Units (incl. Sold)",
                       "Min /sqft (lowest)","Median /sqft (mid)","Max /sqft (highest)",
                       "Min Price","Median Price","Max Price","Avg Unit Price","Total Value (incl. Sold)"]
        topo_show = column_picker(list(tvd.columns), key="topo_cols", locked=["Type"])
        table_with_export(tvd[topo_show], "Topology_View.xlsx", "exp_topo",
                          title="Muraba Veil Topology Summary")


# ── Tab 6: Building View (full floor-by-floor tower elevation) ─────────────────

BUILDING_COLORS = {
    "2 Bedroom": "#5AA9E6", "3 Bedroom - New": "#7FC8F8", "3 Bedroom": "#3C78D8",
    "3 Bedroom Pool": "#22C1C3", "4 Bedroom Pool": "#0E8C8C", "4 Bedroom XL": "#F2A65A",
    "3 Bedroom Duplex": "#8BC34A", "4 Bedroom Duplex": "#4E8C2F", "5 Bedroom Duplex": "#E6B450",
}
# Static crown floors shown at the very top of the tower (amenity, no units)
ROOF_FLOORS = [73, 72, 71, 70]
ROOF_TEXT = "ROOF POOL, RESTAURANT"

def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))

def _abbr(v):
    """Compact money: 19,939,790 → '19.9M'."""
    v = float(v)
    if v >= 1e9: return f"{v/1e9:.2f}B"
    if v >= 1e6: return f"{v/1e6:.1f}M"
    if v >= 1e3: return f"{v/1e3:.0f}K"
    return f"{v:.0f}"

def _heat(v, lo, hi):
    """Map a value to a low→high price heatmap colour (blue → amber → red)."""
    t = 0.5 if hi <= lo else max(0.0, min(1.0, (v - lo) / (hi - lo)))
    stops = [(0.0, (60, 120, 216)), (0.5, (242, 166, 90)), (1.0, (220, 70, 70))]
    for i in range(len(stops) - 1):
        a, ca = stops[i]; b, cb = stops[i + 1]
        if t <= b:
            f = (t - a) / (b - a) if b > a else 0
            r, g, bl = (int(ca[j] + (cb[j] - ca[j]) * f) for j in range(3))
            return f"#{r:02X}{g:02X}{bl:02X}"
    return "#DC4646"

def render_building_view(enhanced=False):
    st.subheader("Building View — Muraba Veil" + (" ✦ Enhanced" if enhanced else ""))
    st.caption("A live elevation of the tower beside a typology dashboard. Lit blocks are available, "
               "dim outlined blocks are sold; hover for instant details. **Click a unit (or a typology in "
               "the panel) to highlight all units of that type** — click again to clear. Scroll the tower; "
               "the KPIs and typology panel track counts, availability and value live.")

    color_mode, avail_only = "Typology", False
    if enhanced:
        cc1, cc2 = st.columns([1.5, 1])
        color_mode = cc1.radio("Colour the tower by", ["Typology", "Price / sqft"],
                               horizontal=True, key="bv2_color")
        avail_only = cc2.toggle("Show available only", value=False, key="bv2_avail")

    bdf = df.copy()
    bdf["_fn"] = pd.to_numeric(bdf["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    bdf["_un"] = pd.to_numeric(bdf["Unit"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    units_by_floor = {int(f): g.sort_values("_un") for f, g in bdf.dropna(subset=["_fn"]).groupby("_fn")}
    floor_nums = [int(f) for f in units_by_floor]
    max_floor = max(floor_nums + list(blocked) + [1]); min_floor = 1
    _ps = pd.to_numeric(bdf["Price_sqft"], errors="coerce")
    PLO, PHI = (float(_ps.min()), float(_ps.max())) if len(_ps) else (0.0, 1.0)

    # ── geometry (tower SVG) ──
    W, TW, TX = 880, 560, 200
    cx = TX + TW / 2
    H_STD, H_TALL, GAP, MULL = 26, 42, 3, 3
    PAD_TOP, CROWN_H, BASE_H, PAD_BOT = 18, 50, 32, 18

    defs = (
        '<defs>'
        '<linearGradient id="sky" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0" stop-color="#0E2747"/><stop offset="1" stop-color="#05080F"/></linearGradient>'
        '<radialGradient id="glow" cx="50%" cy="36%" r="58%">'
        '<stop offset="0" stop-color="#5B8DEF" stop-opacity="0.22"/>'
        '<stop offset="1" stop-color="#5B8DEF" stop-opacity="0"/></radialGradient>'
        '<linearGradient id="crown" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0" stop-color="#F0CE78"/><stop offset="1" stop-color="#9C7A2E"/></linearGradient>'
        '<linearGradient id="podium" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0" stop-color="#16243C"/><stop offset="1" stop-color="#0A1322"/></linearGradient>'
        '<pattern id="amen" width="12" height="12" patternTransform="rotate(45)" patternUnits="userSpaceOnUse">'
        '<rect width="12" height="12" fill="#162844"/><rect width="6" height="12" fill="#21385c"/></pattern>'
        '</defs>')

    body, y, floor_y = [], PAD_TOP + CROWN_H, {}
    # static roof amenity band (levels 70–73)
    roof_h = len(ROOF_FLOORS) * H_STD + (len(ROOF_FLOORS) - 1) * GAP
    body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{roof_h:.0f}" rx="3" fill="url(#amen)"/>')
    body.append(f'<text x="{cx:.0f}" y="{y+roof_h/2+4:.0f}" text-anchor="middle" font-size="12" '
                f'pointer-events="none" fill="#cdd6e2" letter-spacing="2" font-family="Calibri,Arial">{ROOF_TEXT}</text>')
    for j, rf in enumerate(ROOF_FLOORS):
        ry = y + j * (H_STD + GAP) + H_STD / 2
        floor_y[rf] = ry
        body.append(f'<text x="{TX+TW+8:.0f}" y="{ry+3.5:.0f}" text-anchor="start" font-size="10" '
                    f'pointer-events="none" fill="#8aa0bd" font-family="Calibri,Arial">{rf}</text>')
    y += roof_h + GAP

    for f in range(max_floor, min_floor - 1, -1):
        if f in ROOF_FLOORS:
            continue
        g = units_by_floor.get(f)
        is_blocked = f in blocked
        types = list(g["Type"]) if g is not None else []
        tall = g is not None and any(("Pool" in t or "Duplex" in t) for t in types)
        h = H_TALL if tall else H_STD
        floor_y[f] = y + h / 2
        body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="3" fill="#0b1830"/>')
        body.append(f'<text x="{TX+TW+8:.0f}" y="{y+h/2+3.5:.0f}" text-anchor="start" font-size="10" '
                    f'pointer-events="none" fill="#8aa0bd" font-family="Calibri,Arial">{f}</text>')
        if g is not None and len(g):
            n = len(g); cw = (TW - (n - 1) * MULL) / n
            for i, (_, u) in enumerate(g.iterrows()):
                xi = TX + i * (cw + MULL)
                sold = u["Status"] == "Sold"
                col = (_heat(float(u["Price_sqft"]), PLO, PHI)
                       if color_mode == "Price / sqft" else BUILDING_COLORS.get(u["Type"], "#7f8c9b"))
                if avail_only and sold:                       # show available only → ghost the sold cells
                    body.append(f'<rect x="{xi:.1f}" y="{y+1:.0f}" width="{cw:.1f}" height="{h-2}" rx="2.5" '
                                f'fill="#0e1d36" stroke="#16263f"/>')
                    continue
                gattr = (f'class="u" data-u="{_esc(str(u["Unit"]))}" data-ty="{_esc(u["Type"])}" '
                         f'data-fl="{_esc(str(u["Floor"]))}" data-st="{u["Status"]}" '
                         f'data-pr="{_esc(aed(u["Price"]))}" data-ps="{u["Price_sqft"]:,.0f}" '
                         f'data-ai="{u["Internal_sqft"]:,.0f}" data-ae="{u["External_sqft"]:,.0f}" '
                         f'data-at="{u["Total_sqft"]:,.0f}" data-c="{col}"')
                if sold:
                    body.append(f'<g {gattr}><rect x="{xi:.1f}" y="{y+1:.0f}" width="{cw:.1f}" height="{h-2}" '
                                f'rx="2.5" fill="{col}" fill-opacity="0.16" stroke="{col}" stroke-opacity="0.55"/></g>')
                else:
                    hl = max(4, (h - 2) * 0.42)
                    body.append(f'<g {gattr}><rect x="{xi:.1f}" y="{y+1:.0f}" width="{cw:.1f}" height="{h-2}" rx="2.5" fill="{col}"/>'
                                f'<rect x="{xi:.1f}" y="{y+1:.0f}" width="{cw:.1f}" height="{hl:.0f}" rx="2.5" '
                                f'fill="#ffffff" fill-opacity="0.16"/></g>')
                if cw > 30:
                    uabbr = TYPE_ABBR.get(u["Type"], u["Type"])
                    face = f'{_esc(str(u["Unit"]))} &middot; {_esc(uabbr)}'
                if cw > 30 and sold:
                    body.append(f'<text x="{xi+cw/2:.1f}" y="{y+h/2+3:.0f}" text-anchor="middle" font-size="8" '
                                f'pointer-events="none" fill="#ffffff" fill-opacity="0.7" '
                                f'font-family="Calibri,Arial">{face}</text>')
                elif cw > 30:
                    body.append(f'<text x="{xi+cw/2:.1f}" y="{y+h/2-2:.0f}" text-anchor="middle" font-size="8" '
                                f'pointer-events="none" fill="#ffffff" fill-opacity="0.95" '
                                f'font-family="Calibri,Arial">{face}</text>')
                    body.append(f'<text x="{xi+cw/2:.1f}" y="{y+h/2+10:.0f}" text-anchor="middle" font-size="9.5" '
                                f'font-weight="bold" pointer-events="none" fill="#FCE9B0" '
                                f'font-family="Calibri,Arial">{aed(u["Price"])}</text>')
        elif is_blocked:
            body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="3" fill="url(#amen)"/>'
                        f'<text x="{cx:.0f}" y="{y+h/2+3.2:.0f}" text-anchor="middle" font-size="9" '
                        f'pointer-events="none" fill="#9fb3d0" letter-spacing="2" font-family="Calibri,Arial">MEP / MAJLIS</text>')
        else:
            body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="3" fill="none" '
                        f'stroke="#1c2c46" stroke-dasharray="3 3"/>')
        y += h + GAP

    tower_bottom = y - GAP
    total_h = tower_bottom + BASE_H + PAD_BOT

    crown = (
        f'<polygon points="{cx-22:.0f},{PAD_TOP+16} {cx+22:.0f},{PAD_TOP+16} '
        f'{TX+TW:.0f},{PAD_TOP+CROWN_H} {TX:.0f},{PAD_TOP+CROWN_H}" fill="url(#crown)"/>'
        f'<line x1="{cx:.0f}" y1="{PAD_TOP}" x2="{cx:.0f}" y2="{PAD_TOP+16}" stroke="#F0CE78" stroke-width="2"/>'
        f'<circle cx="{cx:.0f}" cy="{PAD_TOP}" r="3" fill="#F0CE78"/>')
    base = (
        f'<rect x="{TX-22:.0f}" y="{tower_bottom+4:.0f}" width="{TW+44}" height="{BASE_H-10}" rx="5" fill="url(#podium)"/>'
        f'<text x="{cx:.0f}" y="{tower_bottom+4+(BASE_H-10)/2+4:.0f}" text-anchor="middle" font-size="11" '
        f'fill="#E6B450" letter-spacing="5" font-family="Calibri,Arial">MURABA VEIL</text>'
        f'<rect x="{TX-50:.0f}" y="{tower_bottom+BASE_H-4:.0f}" width="{TW+100}" height="5" rx="2" fill="#0A1322"/>')
    bg = (f'<rect x="0" y="0" width="{W}" height="{total_h:.0f}" fill="url(#sky)"/>'
          f'<ellipse cx="{cx:.0f}" cy="{total_h*0.4:.0f}" rx="{W*0.55:.0f}" ry="{total_h*0.45:.0f}" fill="url(#glow)"/>')

    # ── left-side level leader-lines (brochure style) ──
    info = {}
    for t in bdf["Type"].unique():
        fls = sorted({int(x) for x in bdf[bdf["Type"] == t]["_fn"].dropna() if int(x) in floor_y})
        if not fls:
            continue
        info[t] = {"mid": sum(floor_y[f] for f in fls) / len(fls), "lo": min(fls), "hi": max(fls)}
    a_types = sorted(info, key=lambda t: info[t]["mid"])
    k = max(1, len(a_types))
    y_top, y_bot, LX = PAD_TOP + CROWN_H + 8, tower_bottom - 8, 192
    ann = []
    for i, t in enumerate(a_types):
        slot = y_top + (i + 0.5) * (y_bot - y_top) / k
        inf = info[t]
        lvl = f"LEVEL {inf['lo']}" if inf["lo"] == inf["hi"] else f"LEVEL {inf['lo']}–{inf['hi']}"
        ann.append(f'<text x="{LX}" y="{slot-3:.0f}" text-anchor="end" font-size="10.5" font-weight="bold" '
                   f'fill="#eef3f9" font-family="Calibri,Arial">{_esc(t.upper())}</text>')
        ann.append(f'<text x="{LX}" y="{slot+9:.0f}" text-anchor="end" font-size="9" fill="#9fb3d0" '
                   f'font-family="Calibri,Arial">{lvl}</text>')
        ann.append(f'<polyline points="{LX+8},{slot:.0f} {(LX+TX)/2:.0f},{slot:.0f} {TX:.0f},{inf["mid"]:.0f}" '
                   f'fill="none" stroke="#6f86a6" stroke-width="1"/>')
        ann.append(f'<circle cx="{TX:.0f}" cy="{inf["mid"]:.0f}" r="3" fill="#eef3f9"/>')

    svg = (f'<svg xmlns="http://www.w3.org/2000/svg" width="{W}" height="{total_h:.0f}" '
           f'viewBox="0 0 {W} {total_h:.0f}">{defs}{bg}{crown}{"".join(body)}{"".join(ann)}{base}</svg>')

    # ── per-typology stats + KPIs (live) ──
    def _stat(t):
        gg = bdf[bdf["Type"] == t]; ta = float(gg["Total_sqft"].sum())
        return {"n": len(gg), "av": int((gg["Status"] == "Available").sum()),
                "so": int((gg["Status"] == "Sold").sum()), "val": float(gg["Price"].sum()),
                "psf": (float(gg["Price"].sum()) / ta) if ta else 0.0}
    present = [t for t in BUILDING_COLORS if t in set(bdf["Type"])]
    present += [t for t in sorted(bdf["Type"].unique()) if t not in present]

    TU = len(df); AV = int((df["Status"] == "Available").sum()); SO = TU - AV
    VAL = float(df["Price"].sum()); AVAL = float(df[df["Status"] == "Available"]["Price"].sum())
    TA = float(df["Total_sqft"].sum()); PSF = VAL / TA if TA else 0.0
    STp = (SO / TU * 100) if TU else 0.0

    def _kpi(label, val, sub=None):
        s = f'<div class="ks">{sub}</div>' if sub else ""
        return f'<div class="kpi"><div class="kl">{label}</div><div class="kv">{val}</div>{s}</div>'
    kpis = "".join([
        _kpi("Total Units", f"{TU}"),
        _kpi("Available", f"{AV}", f"{(AV/TU*100):.0f}% of stock" if TU else ""),
        _kpi("Sold", f"{SO}", f"{STp:.0f}% sold"),
        _kpi("Portfolio Value", aed(VAL)),
        _kpi("Available Value", aed(AVAL)),
        _kpi("Total Sold Value", aed(VAL - AVAL)),
        _kpi("Avg Price/sqft", aed(PSF), "on total area"),
    ])
    rows = []
    for t in present:
        s = _stat(t); col = BUILDING_COLORS.get(t, "#7f8c9b")
        avpct = (s["av"] / s["n"] * 100) if s["n"] else 0
        soldpct = (s["so"] / s["n"] * 100) if s["n"] else 0
        rag = "#5AD18B" if soldpct >= 66 else ("#F0CE78" if soldpct >= 33 else "#F08C8C")
        rag_dot = f'<span class="rag" style="background:{rag}"></span>' if enhanced else ""
        sold_txt = f' &middot; {soldpct:.0f}% sold' if enhanced else ""
        rows.append(
            f'<div class="lg" data-ty="{_esc(t)}"><span class="sw" style="background:{col}"></span>'
            f'<div class="lgm"><div class="lgt">{rag_dot}{_esc(t)}</div>'
            f'<div class="lgs">{s["n"]} units &middot; {s["av"]} avail &middot; {s["so"]} sold{sold_txt}</div>'
            f'</div>'
            f'<div class="lgv">{aed(s["val"])}<div class="lgv2">{aed(s["psf"])} / total sqft</div></div></div>')
    legend_html = "".join(rows)

    # ── enhanced extras: revenue hero + trophies + price scale ──
    hero_html, scale_html = "", ""
    if enhanced:
        SOLDV = VAL - AVAL
        cap = (SOLDV / VAL * 100) if VAL else 0
        av_df = df[df["Status"] == "Available"]
        if len(av_df):
            top = av_df.loc[av_df["Price"].idxmax()]
            top_txt = f'{_esc(str(top["Unit"]))} · {_esc(top["Type"])} · {aed(top["Price"])}'
        else:
            top_txt = "—"
        ph = df[df["Type"] == "5 Bedroom Duplex"]
        ph_txt = ph.iloc[0]["Status"] if len(ph) else "—"
        ph_col = "#5AD18B" if ph_txt == "Available" else ("#F08C8C" if ph_txt == "Sold" else "#9fb3d0")
        hero_html = (
            '<div class="hero">'
            '<div class="ht">Revenue captured vs. remaining</div>'
            f'<div class="hbar"><div class="hb-f" style="width:{cap:.1f}%"></div>'
            f'<div class="hb-r" style="width:{100-cap:.1f}%"></div></div>'
            f'<div class="hl"><span><b style="color:#5AD18B">{aed(SOLDV)}</b> sold · {cap:.0f}%</span>'
            f'<span><b style="color:#F0CE78">{aed(AVAL)}</b> available · {100-cap:.0f}%</span></div>'
            f'<div class="htro">🏆 Top available: <b>{top_txt}</b><br>👑 Penthouse: '
            f'<b style="color:{ph_col}">{ph_txt}</b></div></div>')
        if color_mode == "Price / sqft":
            scale_html = (f'<div class="scale"><span>AED {PLO:,.0f}</span>'
                          f'<span class="grad"></span><span>AED {PHI:,.0f}</span></div>')

    css = """<style>
    *{box-sizing:border-box;}
    .bv{display:flex;gap:14px;height:720px;background:#05080F;border-radius:12px;padding:12px;
        font-family:Calibri,Arial,sans-serif;}
    .tower{flex:0 0 920px;overflow-y:auto;border-radius:10px;}
    .tower svg{display:block;margin:0 auto;}
    .tower::-webkit-scrollbar,.legend::-webkit-scrollbar{width:8px;}
    .tower::-webkit-scrollbar-thumb,.legend::-webkit-scrollbar-thumb{background:#24395c;border-radius:5px;}
    .side{flex:1;min-width:230px;max-width:360px;display:flex;flex-direction:column;gap:12px;color:#e6edf6;}
    .kpis{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;}
    .kpi{background:#0d1b30;border:1px solid #1e3357;border-radius:9px;padding:9px 11px;}
    .kl{font-size:11px;color:#9fb3d0;}
    .kv{font-size:17px;font-weight:700;margin-top:2px;}
    .ks{font-size:10px;color:#7f97b5;margin-top:1px;}
    .sh{font-size:13px;font-weight:700;color:#cdd6e2;letter-spacing:.4px;}
    .sh span{font-weight:400;color:#7f97b5;font-size:11px;}
    .legend{overflow-y:auto;display:flex;flex-direction:column;gap:7px;padding-right:5px;}
    .lg{display:flex;align-items:center;gap:10px;background:#0b1830;border:1px solid #16263f;
        border-radius:9px;padding:8px 11px;}
    .sw{width:14px;height:14px;border-radius:4px;flex:none;}
    .lgm{flex:1;min-width:0;}
    .lgt{font-size:13px;font-weight:600;}
    .lgs{font-size:11px;color:#9fb3d0;margin:1px 0 5px;}
    .bar{height:5px;background:#1c2c46;border-radius:3px;overflow:hidden;}
    .bar span{display:block;height:100%;border-radius:3px;}
    .lgv{text-align:right;font-size:13px;font-weight:700;color:#F0CE78;white-space:nowrap;}
    .lgv2{font-size:10px;color:#7f97b5;font-weight:400;}
    .lg{cursor:pointer;transition:border-color .12s;}
    .lg.on{border-color:#F0CE78;background:#13233d;}
    .u{cursor:pointer;}
    .u.dim{opacity:.14;}
    .u.hot rect{stroke:#ffffff;stroke-width:1.6;}
    #bvtip{position:fixed;display:none;z-index:99999;pointer-events:none;background:#0d1b30;
        border:1px solid #2a3f5f;border-left:4px solid #888;border-radius:9px;padding:9px 12px;
        color:#e6edf6;font:12px/1.55 Calibri,Arial;box-shadow:0 8px 28px rgba(0,0,0,.55);max-width:300px;}
    #bvtip .h{font-size:13px;font-weight:700;margin-bottom:3px;}
    #bvtip .p{color:#F0CE78;font-weight:700;margin-top:4px;}
    #bvtip .a{color:#9fb3d0;margin-top:3px;font-size:11px;}
    .rag{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:6px;vertical-align:middle;}
    .hero{background:#0d1b30;border:1px solid #1e3357;border-radius:10px;padding:10px 12px;}
    .ht{font-size:11px;color:#9fb3d0;margin-bottom:6px;letter-spacing:.3px;}
    .hbar{display:flex;height:13px;border-radius:7px;overflow:hidden;background:#16263f;}
    .hb-f{background:linear-gradient(90deg,#2f8f5d,#5AD18B);}
    .hb-r{background:linear-gradient(90deg,#3a5078,#24395c);}
    .hl{display:flex;justify-content:space-between;font-size:11.5px;margin-top:6px;color:#cdd6e2;}
    .htro{font-size:11px;color:#9fb3d0;margin-top:7px;line-height:1.6;}
    .scale{display:flex;align-items:center;gap:8px;font-size:10px;color:#9fb3d0;margin:-2px 2px 0;}
    .scale .grad{flex:1;height:8px;border-radius:4px;background:linear-gradient(90deg,#3C78D8,#F2A65A,#DC4646);}
    </style>"""
    dyn = (f'<div class="bv"><div class="tower">{svg}</div>'
           f'<div class="side">{hero_html}<div class="kpis">{kpis}</div>'
           f'<div class="sh">By Typology <span>&nbsp;live counts &amp; value</span></div>'
           f'{scale_html}'
           f'<div class="legend">{legend_html}</div></div></div><div id="bvtip"></div>')
    js = """<script>
    (function(){
      var tip=document.getElementById('bvtip'), tw=document.querySelector('.tower');
      var units=Array.prototype.slice.call(document.querySelectorAll('.u'));
      var rows=Array.prototype.slice.call(document.querySelectorAll('.lg'));
      var focus=null;
      function applyFocus(){
        units.forEach(function(el){
          el.classList.remove('hot','dim');
          if(focus){ el.classList.add(el.dataset.ty===focus?'hot':'dim'); }
        });
        rows.forEach(function(r){ r.classList.toggle('on', !!focus && r.dataset.ty===focus); });
      }
      function setFocus(t){ focus=(t===focus)?null:t; applyFocus(); }
      function move(e){
        var el=e.target.closest('.u');
        if(!el){tip.style.display='none';return;}
        var d=el.dataset, sc=(d.st==='Sold')?'#F08C8C':'#5AD18B';
        tip.style.borderLeftColor=d.c;
        tip.innerHTML='<div class="h">'+d.u+' &middot; '+d.ty+'</div>'+
          '<div>Floor '+d.fl+' &middot; <b style="color:'+sc+'">'+d.st+'</b></div>'+
          '<div class="p">'+d.pr+' <span style="font-weight:400;color:#9fb3d0">('+d.ps+'/sqft)</span></div>'+
          '<div class="a">Internal '+d.ai+' &middot; External '+d.ae+' &middot; Total '+d.at+' sqft</div>';
        tip.style.display='block';
        var x=e.clientX+16, y=e.clientY+16, tw2=tip.offsetWidth, th=tip.offsetHeight;
        if(x+tw2>window.innerWidth-8)  x=e.clientX-tw2-16;
        if(y+th>window.innerHeight-8)  y=e.clientY-th-16;
        tip.style.left=x+'px'; tip.style.top=y+'px';
      }
      if(tw){
        tw.addEventListener('mousemove',move);
        tw.addEventListener('mouseleave',function(){tip.style.display='none';});
        tw.addEventListener('click',function(e){ var el=e.target.closest('.u'); setFocus(el?el.dataset.ty:null); });
      }
      rows.forEach(function(r){ r.addEventListener('click',function(){ setFocus(r.dataset.ty); }); });
    })();
    </script>"""
    # data signature → the iframe srcdoc changes on any edit, forcing a live refresh
    sig = (f"{len(df)}|{int(df['Price'].sum())}|{int((df['Status']=='Sold').sum())}|"
           f"{color_mode}|{int(avail_only)}")
    components.html(f"<!--bv:{_esc(sig)}-->" + css + dyn + js, height=748, scrolling=False)


BROCHURE_COLORS = {
    "2 Bedroom": "#6E7E6A", "3 Bedroom - New": "#7E8C77", "3 Bedroom": "#566A54",
    "3 Bedroom Pool": "#5C7A7C", "4 Bedroom Pool": "#46625F", "4 Bedroom XL": "#A8743C",
    "3 Bedroom Duplex": "#7A7A3C", "4 Bedroom Duplex": "#5C5E2E", "5 Bedroom Duplex": "#9C7A3A",
}

def render_building_view_brochure():
    st.caption("The live tower in the brochure's warm architectural palette — tan page, sepia "
               "line-work, muted earth tones. Filled units are available, hollow units are sold; hover "
               "for details, click a unit or typology to focus. Level leader-lines as in the plan.")

    bdf = df.copy()
    bdf["_fn"] = pd.to_numeric(bdf["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    bdf["_un"] = pd.to_numeric(bdf["Unit"].astype(str).str.replace(r"[^0-9]", "", regex=True), errors="coerce")
    units_by_floor = {int(f): g.sort_values("_un") for f, g in bdf.dropna(subset=["_fn"]).groupby("_fn")}
    floor_nums = [int(f) for f in units_by_floor]

    def _is_up(u):                                      # every duplex occupies its floor + the one ABOVE
        v = u.get("Dup_Up")
        return bool(v) if pd.notna(v) else False

    # duplexes also occupy the floor above their record floor → counts toward building height
    up_tops = [int(f) + 1 for f, g in units_by_floor.items()
               for _, u in g.iterrows() if "Duplex" in str(u["Type"]) and _is_up(u)]
    max_floor = max(floor_nums + list(blocked) + up_tops + [1]); min_floor = 1
    ROOF_N = 4                                          # the top 4 floors are always ROOF POOL, RESTAURANT
    roof_floors = [max_floor + i for i in range(ROOF_N, 0, -1)]   # e.g. 69 -> [73,72,71,70]

    # ── brochure palette ──
    PAGE, INK, SUB = "#B4A48D", "#33302A", "#6E6657"
    SLAB, SLABLN, SOLDS = "#AC9C84", "#9C8E76", "#8C8270"

    W, TW, TX = 880, 560, 200
    cx = TX + TW / 2
    H_STD, H_TALL, GAP, MULL = 26, 42, 3, 3
    PAD_TOP, CROWN_H, BASE_H, PAD_BOT = 18, 50, 30, 18

    defs = ('<defs><pattern id="amenP" width="10" height="10" patternTransform="rotate(45)" '
            'patternUnits="userSpaceOnUse"><rect width="10" height="10" fill="#A2937B"/>'
            '<rect width="5" height="10" fill="#94866F"/></pattern></defs>')

    body, y, floor_y, floor_h = [], PAD_TOP + CROWN_H, {}, {}
    # the top ROOF_N floors are always the roof amenity band (moves up as floors are added)
    roof_h = len(roof_floors) * H_STD + (len(roof_floors) - 1) * GAP
    roof_bottom = y + roof_h                        # bottom edge of the roof band (y here = its top)
    body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{roof_h:.0f}" rx="2" fill="url(#amenP)"/>')
    body.append(f'<text x="{cx:.0f}" y="{y+roof_h/2+4.5:.0f}" text-anchor="middle" font-size="13" '
                f'font-weight="bold" pointer-events="none" fill="{INK}" letter-spacing="2.5" '
                f'font-family="Calibri,Arial">{ROOF_TEXT}</text>')
    for j, rf in enumerate(roof_floors):
        ry = y + j * (H_STD + GAP) + H_STD / 2
        floor_y[rf] = ry; floor_h[rf] = H_STD
        body.append(f'<text x="{TX+TW+8:.0f}" y="{ry+3.5:.0f}" text-anchor="start" font-size="10" '
                    f'pointer-events="none" fill="{SUB}" font-family="Calibri,Arial">{rf}</text>')
    y += roof_h + GAP

    dup_units = []                      # duplexes drawn after the loop, spanning two floors

    def draw_box(xi, cw, ry, rh, ty_mid, sold, col, u, span=None):
        gfl = span if span else str(u["Floor"])
        _tot = float(u["Total_sqft"]) if pd.notna(u["Total_sqft"]) else 0.0
        _pt = (float(u["Price"]) / _tot) if _tot else 0.0          # price per TOTAL sqft (for the tooltip)
        _av = pd.to_numeric(u.get("Adj_Pct"), errors="coerce")     # appreciation(+)/discount(-) %, for the tooltip
        _adj = "" if (pd.isna(_av) or float(_av) == 0) else (("+" if float(_av) > 0 else "−") + f"{abs(float(_av)):.1f}%")
        gattr = (f'class="u" data-u="{_esc(str(u["Unit"]))}" data-ty="{_esc(u["Type"])}" '
                 f'data-fl="{_esc(gfl)}" data-st="{u["Status"]}" '
                 f'data-pr="{_esc(aed(u["Price"]))}" data-ps="{u["Price_sqft"]:,.0f}" data-pt="{_pt:,.0f}" '
                 f'data-ai="{u["Internal_sqft"]:,.0f}" data-ae="{u["External_sqft"]:,.0f}" '
                 f'data-at="{u["Total_sqft"]:,.0f}" data-adj="{_esc(_adj)}" data-c="{col}"')
        if sold:
            body.append(f'<g {gattr}><rect x="{xi:.1f}" y="{ry:.1f}" width="{cw:.1f}" height="{rh:.1f}" '
                        f'rx="2" fill="none" stroke="{SOLDS}" stroke-width="1"/></g>')
        else:
            body.append(f'<g {gattr}><rect x="{xi:.1f}" y="{ry:.1f}" width="{cw:.1f}" height="{rh:.1f}" '
                        f'rx="2" fill="{col}" stroke="{INK}" stroke-width="0.5"/></g>')
        if cw > 30:
            uabbr = TYPE_ABBR.get(u["Type"], u["Type"])
            face = f'{_esc(str(u["Unit"]))} &middot; {_esc(uabbr)}'
            tcol = SUB if sold else "#EDE6D7"
            body.append(f'<text x="{xi+cw/2:.1f}" y="{ty_mid-2:.1f}" text-anchor="middle" font-size="8" '
                        f'pointer-events="none" fill="{tcol}" font-family="Calibri,Arial">{face}</text>')
            if sold:
                body.append(f'<text x="{xi+cw/2:.1f}" y="{ty_mid+10:.1f}" text-anchor="middle" font-size="9" '
                            f'font-weight="bold" pointer-events="none" fill="{SOLDS}" letter-spacing="1.5" '
                            f'font-family="Calibri,Arial">SOLD</text>')
            else:
                body.append(f'<text x="{xi+cw/2:.1f}" y="{ty_mid+10:.1f}" text-anchor="middle" font-size="9.5" '
                            f'font-weight="bold" pointer-events="none" fill="#FBF3DF" '
                            f'font-family="Calibri,Arial">{aed(u["Price"])}</text>')

    for f in range(max_floor, min_floor - 1, -1):
        if f in roof_floors:
            continue
        g = units_by_floor.get(f)
        is_blocked = f in blocked
        # draw EVERY level so the floor numbering stays continuous. An empty level is either the
        # lower half of a duplex that spans onto it (the duplex box in the post-pass covers it) or a
        # genuinely vacant floor — both render as a plain slab carrying their own floor number.
        types = list(g["Type"]) if g is not None else []
        tall = g is not None and any(("Pool" in t or "Duplex" in t) for t in types)
        h = H_TALL if tall else H_STD
        floor_y[f] = y + h / 2
        floor_h[f] = h
        body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="2" '
                    f'fill="{SLAB}" stroke="{SLABLN}" stroke-width="0.5"/>')
        body.append(f'<text x="{TX+TW+8:.0f}" y="{y+h/2+3.5:.0f}" text-anchor="start" font-size="10" '
                    f'pointer-events="none" fill="{SUB}" font-family="Calibri,Arial">{f}</text>')
        if g is not None and len(g):
            n = len(g); cw = (TW - (n - 1) * MULL) / n
            for i, (_, u) in enumerate(g.iterrows()):
                xi = TX + i * (cw + MULL)
                sold = u["Status"] == "Sold"
                col = BROCHURE_COLORS.get(u["Type"], "#7d7461")
                if "Duplex" in str(u["Type"]):          # 2-floor unit → drawn in post-pass below
                    dup_units.append((f, xi, cw, sold, col, u))
                    continue
                draw_box(xi, cw, y + 1, h - 2, y + h / 2, sold, col, u)
        elif is_blocked:
            lbl = "MAJLIS" if "MAJ" in str(blocked.get(f, "")).upper() else "MEP"
            body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="2" fill="url(#amenP)"/>'
                        f'<text x="{cx:.0f}" y="{y+h/2+4:.0f}" text-anchor="middle" font-size="12" '
                        f'font-weight="bold" pointer-events="none" fill="{INK}" letter-spacing="3" '
                        f'font-family="Calibri,Arial">{lbl}</text>')
        elif f == 2:
            body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="2" fill="url(#amenP)"/>'
                        f'<text x="{cx:.0f}" y="{y+h/2+4:.0f}" text-anchor="middle" font-size="12" '
                        f'font-weight="bold" pointer-events="none" fill="{INK}" letter-spacing="3" '
                        f'font-family="Calibri,Arial">AMENITIES</text>')
        else:                                            # an up-duplex's reserved upper level (covered below)
            body.append(f'<rect x="{TX:.0f}" y="{y:.0f}" width="{TW}" height="{h}" rx="2" fill="none" '
                        f'stroke="{SLABLN}" stroke-dasharray="3 3"/>')
        y += h + GAP

    tower_bottom = y - GAP

    # duplexes: one unit two floors tall. Every duplex is recorded on its LOWER floor and spans UP
    # onto the floor above (e.g. 3BR Duplex 5901 on 59 → 59–60), so it is labelled by its lower floor
    # and the empty reserved level above is covered by the box.
    for (f, xi, cw, sold, col, u) in dup_units:
        lo, hi = (f, f + 1) if _is_up(u) else (f - 1, f)
        top = (floor_y[hi] - floor_h[hi] / 2) if hi in floor_y else (floor_y[lo] - floor_h[lo] / 2 - GAP - H_STD)
        bot = (floor_y[lo] + floor_h[lo] / 2) if lo in floor_y else (floor_y[hi] + floor_h[hi] / 2 + GAP + H_STD)
        top = max(top, roof_bottom + 1)
        ry, rh = top + 1, bot - top - 2
        ty_mid = (top + bot) / 2
        draw_box(xi, cw, ry, rh, ty_mid, sold, col, u, span=f"{lo}–{hi}")
    total_h = tower_bottom + BASE_H + PAD_BOT

    crown = (f'<polygon points="{cx-22:.0f},{PAD_TOP+16} {cx+22:.0f},{PAD_TOP+16} '
             f'{TX+TW:.0f},{PAD_TOP+CROWN_H} {TX:.0f},{PAD_TOP+CROWN_H}" fill="{SLAB}" '
             f'stroke="{INK}" stroke-width="1"/>'
             f'<line x1="{cx:.0f}" y1="{PAD_TOP}" x2="{cx:.0f}" y2="{PAD_TOP+16}" stroke="{INK}" stroke-width="1.5"/>'
             f'<circle cx="{cx:.0f}" cy="{PAD_TOP}" r="3" fill="{INK}"/>')
    base = (f'<rect x="{TX-22:.0f}" y="{tower_bottom+4:.0f}" width="{TW+44}" height="{BASE_H-10}" rx="3" '
            f'fill="{SLAB}" stroke="{SLABLN}"/>'
            f'<text x="{cx:.0f}" y="{tower_bottom+4+(BASE_H-10)/2+4:.0f}" text-anchor="middle" font-size="11" '
            f'fill="{INK}" letter-spacing="5" font-family="Calibri,Arial">MURABA VEIL</text>'
            f'<rect x="{TX-50:.0f}" y="{tower_bottom+BASE_H-4:.0f}" width="{TW+100}" height="4" rx="2" fill="{INK}"/>')
    bg = f'<rect x="0" y="0" width="{W}" height="{total_h:.0f}" fill="{PAGE}"/>'

    ann = ""                                         # typology leader labels hidden (per request)

    svg = (f'<svg xmlns="http://www.w3.org/2000/svg" width="{W}" height="{total_h:.0f}" '
           f'viewBox="0 0 {W} {total_h:.0f}">{defs}{bg}{crown}{"".join(body)}{ann}{base}</svg>')

    def _stat(t):
        gg = bdf[bdf["Type"] == t]; ta = float(gg["Total_sqft"].sum())
        return {"n": len(gg), "av": int((gg["Status"] == "Available").sum()),
                "so": int((gg["Status"] == "Sold").sum()), "val": float(gg["Price"].sum()),
                "psf": (float(gg["Price"].sum()) / ta) if ta else 0.0}
    present = [t for t in BROCHURE_COLORS if t in set(bdf["Type"])]
    present += [t for t in sorted(bdf["Type"].unique()) if t not in present]
    TU = len(df); AV = int((df["Status"] == "Available").sum()); SO = TU - AV
    VAL = float(df["Price"].sum()); AVAL = float(df[df["Status"] == "Available"]["Price"].sum())
    TA = float(df["Total_sqft"].sum()); PSF = VAL / TA if TA else 0.0; STp = (SO / TU * 100) if TU else 0.0

    def _kpi(label, val, sub=None):
        s = f'<div class="ks">{sub}</div>' if sub else ""
        return f'<div class="kpi"><div class="kl">{label}</div><div class="kv">{val}</div>{s}</div>'
    av_pct = (AV / TU * 100) if TU else 0.0
    so_pct = (SO / TU * 100) if TU else 0.0
    SOLDVAL = VAL - AVAL
    sv_pct = (SOLDVAL / VAL * 100) if VAL else 0.0
    avv_pct = (AVAL / VAL * 100) if VAL else 0.0
    ALLOW = ALLOWABLE_SELLABLE                        # fixed design cap (818,187), same as the top header
    BASE_N = base_unit_count()                        # units in the saved Base Version (for the delta card)

    def _m(v):                                       # compact AED for the split-bar labels
        return (f"AED {v/1e9:.2f}B" if v >= 1e9 else
                f"AED {v/1e6:.0f}M" if v >= 1e6 else aed(v))

    def _split(title, lab1, v1, p1, lab2, v2, p2):
        return ('<div class="kpi">'
                f'<div class="kl">{title}</div>'
                f'<div class="split"><div class="seg so" style="width:{p1:.0f}%"></div>'
                f'<div class="seg av" style="width:{p2:.0f}%"></div></div>'
                '<div class="srows">'
                f'<div class="srow"><span class="dot so"></span><span class="sk">{lab1}</span>'
                f'<span class="sv">{v1} &middot; {p1:.0f}%</span></div>'
                f'<div class="srow"><span class="dot av"></span><span class="sk">{lab2}</span>'
                f'<span class="sv">{v2} &middot; {p2:.0f}%</span></div>'
                '</div></div>')

    split_units = _split("Sold vs Available (stock)", "Sold", f"<b>{SO}</b>", so_pct,
                         "Available", f"<b>{AV}</b>", av_pct)
    split_value = _split("Sold vs Available (value)", "Sold", f"<b>{_m(SOLDVAL)}</b>", sv_pct,
                         "Available", f"<b>{_m(AVAL)}</b>", avv_pct)

    add_card = _kpi("Additional from base", f"{TU - BASE_N:+d}", f"base = {BASE_N} units") if BASE_N else ""
    kpis = "".join([
        _kpi("Total Units", f"{TU}"),
        _kpi("Available Stock", f"{AV}"),
        _kpi("Sold", f"{SO}"),
        _kpi("Total Project Value", aed(VAL)),
        _kpi("Available Stock Value", aed(AVAL)),
        _kpi("Total Sold Value", aed(SOLDVAL)),
        _kpi("Avg Price/sqft", aed(PSF), "on total area"),
        _kpi("Total Area (sqft)", f"{TA:,.0f}"),
        _kpi("Total Allowable Sellable (sqft)", f"{ALLOW:,.0f}"),
        add_card,
        split_units, split_value])
    rows = []
    for t in present:
        s = _stat(t); col = BROCHURE_COLORS.get(t, "#7d7461")
        rows.append(
            f'<div class="lg" data-ty="{_esc(t)}"><span class="sw" style="background:{col}"></span>'
            f'<div class="lgm"><div class="lgt">{_esc(t)}</div>'
            f'<div class="lgs">{s["n"]} units &middot; {s["av"]} avail &middot; {s["so"]} sold</div>'
            f'</div>'
            f'<div class="lgv">{aed(s["val"])}<div class="lgv2">{aed(s["psf"])} / total sqft</div></div></div>')
    legend_html = "".join(rows)

    css = """<style>
    *{box-sizing:border-box;}
    .bv{display:flex;gap:14px;height:720px;background:#B4A48D;border-radius:12px;padding:12px;
        font-family:Calibri,Arial,sans-serif;}
    .tower{flex:1 1 560px;min-width:0;overflow:auto;border-radius:10px;}
    .tower svg{display:block;margin:0 auto;width:100%;height:auto;max-width:920px;}
    .tower::-webkit-scrollbar,.legend::-webkit-scrollbar{width:8px;height:8px;}
    .tower::-webkit-scrollbar-thumb,.legend::-webkit-scrollbar-thumb{background:#8C7E66;border-radius:5px;}
    .side{flex:1 1 280px;min-width:240px;max-width:420px;display:flex;flex-direction:column;gap:12px;
        color:#33302A;overflow:hidden;}
    .kpiwrap{flex:1 1 auto;min-height:90px;overflow-y:auto;padding-right:3px;}
    /* narrow screens: stack the panel under the tower so nothing is clipped */
    @media (max-width:900px){
      .bv{flex-direction:column;height:auto;}
      .tower{flex:1 1 auto;}
      .side{max-width:none;overflow:visible;}
      .kpiwrap,.legend{flex:none;height:auto;min-height:0;max-height:none;overflow:visible;}
    }
    .kpis{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;}
    .kpi{background:linear-gradient(180deg,#CCC0AB,#C2B49C);border:1px solid #9C8E76;border-radius:11px;
         padding:11px 13px;display:flex;flex-direction:column;justify-content:flex-start;gap:3px;
         min-height:80px;box-shadow:0 1px 2px rgba(60,50,30,.10);}
    .kl{font-size:10px;color:#6E6657;letter-spacing:.5px;text-transform:uppercase;font-weight:700;
        line-height:1.25;min-height:25px;}
    .kv{font-size:17px;font-weight:800;color:#2E2B25;line-height:1.15;white-space:nowrap;letter-spacing:-.2px;}
    .ks{font-size:10px;color:#7d735f;}
    .split{display:flex;height:12px;border-radius:6px;overflow:hidden;margin:9px 0 8px;background:#A2937B;}
    .split .seg{height:100%;}
    .split .seg.so{background:#B5532F;}
    .split .seg.av{background:#3E7A4E;}
    .srows{display:flex;flex-direction:column;gap:4px;}
    .srow{display:flex;align-items:center;gap:7px;font-size:11px;color:#5b5346;white-space:nowrap;}
    .srow .sk{color:#6E6657;}
    .srow .sv{margin-left:auto;color:#2E2B25;}
    .srow .sv b{font-size:13px;font-weight:800;}
    .dot{display:inline-block;width:9px;height:9px;border-radius:50%;flex:none;}
    .dot.so{background:#B5532F;}
    .dot.av{background:#3E7A4E;}
    .sh{font-size:13px;font-weight:700;color:#33302A;letter-spacing:.4px;}
    .sh span{font-weight:400;color:#7d735f;font-size:11px;}
    .legend{flex:0 0 244px;min-height:0;overflow-y:auto;display:flex;flex-direction:column;gap:7px;padding-right:5px;}
    .lg{display:flex;align-items:center;gap:10px;background:#C6B9A4;border:1px solid #9C8E76;
        border-radius:9px;padding:8px 11px;cursor:pointer;transition:border-color .12s;}
    .lg.on{border-color:#33302A;background:#CFC2AD;}
    .sw{width:14px;height:14px;border-radius:4px;flex:none;}
    .lgm{flex:1;min-width:0;}
    .lgt{font-size:13px;font-weight:600;}
    .lgs{font-size:11px;color:#6E6657;margin-top:1px;}
    .lgv{text-align:right;font-size:13px;font-weight:700;color:#5C4A28;white-space:nowrap;}
    .lgv2{font-size:10px;color:#7d735f;font-weight:400;}
    .u{cursor:pointer;}
    .u.dim{opacity:.16;}
    .u.hot rect{stroke:#33302A;stroke-width:1.8;}
    #bvtip2{position:fixed;display:none;z-index:99999;pointer-events:none;background:#F3ECDD;
        border:1px solid #B8A98E;border-left:4px solid #8A6D3B;border-radius:9px;padding:9px 12px;
        color:#33302A;font:12px/1.55 Calibri,Arial;box-shadow:0 8px 22px rgba(60,50,30,.35);max-width:300px;}
    #bvtip2 .h{font-size:13px;font-weight:700;margin-bottom:3px;}
    #bvtip2 .p{color:#5C4A28;font-weight:700;margin-top:4px;}
    #bvtip2 .a{color:#6E6657;margin-top:3px;font-size:11px;}
    #printbtn{background:#8A6D3B;color:#F3ECDD;border:none;border-radius:9px;padding:10px 12px;
        font:700 12px Calibri,Arial;cursor:pointer;letter-spacing:.5px;width:100%;}
    #printbtn:hover{background:#755C32;}
    /* Print / Save-as-PDF: the on-screen side-by-side view (tower + stats) on ONE A1 sheet.
       Deterministic: the .bv is a FIXED width and the tower SVG keeps its natural size (capped so it
       can't blow up), so the layout doesn't depend on the iframe width and never paginates. */
    @media print{
      @page{size:A1 portrait;margin:9mm;}
      html,body{background:#fff !important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
      #printbtn{display:none !important;}
      .bv{display:flex !important;flex-direction:row !important;align-items:flex-start !important;
          width:560mm !important;max-width:560mm !important;margin:0 auto !important;height:auto !important;
          overflow:visible !important;gap:7mm !important;padding:6mm !important;border-radius:0 !important;
          page-break-inside:avoid;break-inside:avoid;}
      .tower{flex:0 0 auto !important;overflow:visible !important;}
      .tower svg{width:auto !important;height:auto !important;max-width:250mm !important;max-height:775mm !important;}
      .side{flex:1 1 auto !important;min-width:0 !important;max-width:none !important;overflow:visible !important;}
      .kpiwrap,.legend{flex:none !important;height:auto !important;max-height:none !important;overflow:visible !important;}
      .kpis{grid-template-columns:repeat(2,1fr) !important;}
      #bvtip2{display:none !important;}
    }
    </style>"""
    dyn = (f'<div class="bv"><div class="tower">{svg}</div>'
           f'<div class="side"><button id="printbtn" onclick="window.print()">'
           f'&#128424;&nbsp; Print / Save as PDF (A1)</button>'
           f'<div class="kpiwrap"><div class="kpis">{kpis}</div></div>'
           f'<div class="sh">By Typology <span>&nbsp;live counts &amp; value</span></div>'
           f'<div class="legend">{legend_html}</div></div></div><div id="bvtip2"></div>')
    js = """<script>
    (function(){
      var tip=document.getElementById('bvtip2'), tw=document.querySelector('.tower');
      var units=Array.prototype.slice.call(document.querySelectorAll('.u'));
      var rows=Array.prototype.slice.call(document.querySelectorAll('.lg'));
      var focus=null;
      function applyFocus(){
        units.forEach(function(el){ el.classList.remove('hot','dim');
          if(focus){ el.classList.add(el.dataset.ty===focus?'hot':'dim'); } });
        rows.forEach(function(r){ r.classList.toggle('on', !!focus && r.dataset.ty===focus); });
      }
      function setFocus(t){ focus=(t===focus)?null:t; applyFocus(); }
      function move(e){
        var el=e.target.closest('.u');
        if(!el){tip.style.display='none';return;}
        var d=el.dataset, sc=(d.st==='Sold')?'#B5532F':'#3E7A4E';
        tip.style.borderLeftColor=d.c;
        tip.innerHTML='<div class="h">'+d.u+' &middot; '+d.ty+'</div>'+
          '<div>Floor '+d.fl+' &middot; <b style="color:'+sc+'">'+d.st+'</b></div>'+
          '<div class="p">'+d.pr+' <span style="font-weight:400;color:#6E6657">(AED '+d.pt+' / total sqft)</span></div>'+
          (d.adj?('<div style="font-weight:700;color:'+(d.adj.charAt(0)==='+'?'#1a7f37':'#d1242f')+'">'
                  +(d.adj.charAt(0)==='+'?'Appreciation ':'Discount ')+d.adj+'</div>'):'')+
          '<div class="a">Internal '+d.ai+' &middot; External '+d.ae+' &middot; Total '+d.at+' sqft</div>';
        tip.style.display='block';
        var x=e.clientX+16, y=e.clientY+16, tw2=tip.offsetWidth, th=tip.offsetHeight;
        if(x+tw2>window.innerWidth-8)  x=e.clientX-tw2-16;
        if(y+th>window.innerHeight-8)  y=e.clientY-th-16;
        tip.style.left=x+'px'; tip.style.top=y+'px';
      }
      if(tw){ tw.addEventListener('mousemove',move);
              tw.addEventListener('mouseleave',function(){tip.style.display='none';});
              tw.addEventListener('click',function(e){ var el=e.target.closest('.u'); setFocus(el?el.dataset.ty:null); }); }
      rows.forEach(function(r){ r.addEventListener('click',function(){ setFocus(r.dataset.ty); }); });
    })();
    </script>"""
    sig = f"{len(df)}|{int(df['Price'].sum())}|{int((df['Status']=='Sold').sum())}|plan"
    components.html(f"<!--bvp:{_esc(sig)}-->" + css + dyn + js, height=748, scrolling=True)


if tab6 is not None:                 # only when enabled (hidden on the published app)
    with tab6:
        render_building_view(enhanced=False)
if tab6b is not None:
    with tab6b:
        render_building_view(enhanced=True)
if tab6c is not None:
    with tab6c:
        render_building_view_brochure()


# ── Tab 3: Floor Manager ───────────────────────────────────────────────────────

with tab3:
    floors = st.session_state.floors

    # Parameters
    def terrace_group_key(t):
        if t in ("2 Bedroom", "3 Bedroom - New", "3 Bedroom"): return "standard"
        if t == "3 Bedroom Pool":                          return "3 Bedroom Pool"
        if t == "4 Bedroom Pool":                          return "4 Bedroom Pool"
        if t == "4 Bedroom XL":                            return "simplex"
        if t in ("3 Bedroom Duplex", "4 Bedroom Duplex"):  return "duplex"
        return None   # 5 Bedroom Duplex → fixed 100%

    with st.expander("⚙️  Escalation & Terrace Settings (per typology, from launch sheets)", expanded=True):
        st.caption("Pick a typology, then set its **escalation** (price/sqft added per floor as you go up) "
                   "and **terrace %**. These are the building-wide defaults used when pricing new or "
                   "edited floors. The table below shows the current settings for every typology.")
        s_type = st.selectbox("Typology", UNIT_TYPES, key="set_type")

        sc1, sc2, sc3 = st.columns(3)
        cur_esc = float(params["escalation"].get(s_type, 0.0))
        new_e = sc1.number_input("Escalation (AED/sqft per floor)", min_value=0.0, step=1.0,
                                 value=cur_esc, key=f"set_esc_{s_type}")

        gkey = terrace_group_key(s_type)
        if gkey is not None:
            cur_tr = float(params["terrace"][gkey]) * 100
            new_t = sc2.number_input("Terrace %", min_value=0.0, max_value=100.0, step=1.0,
                                     value=cur_tr, key=f"set_tr_{s_type}")
        else:
            sc2.number_input("Terrace %", min_value=0.0, max_value=100.0, step=1.0, value=100.0,
                             key=f"set_tr_{s_type}", disabled=True,
                             help="5 Bedroom Duplex (penthouse) terrace is fixed at 100%.")
            new_t = None

        cur_dpx = float(params.get("duplex_premium", 0.0))
        if "Duplex" in s_type:
            new_dpx = sc3.number_input("Duplex Premium (AED/sqft)", min_value=0.0, step=50.0,
                                       value=cur_dpx, key="set_dpx")
        else:
            sc3.number_input("Duplex Premium (AED/sqft)", min_value=0.0, step=50.0, value=cur_dpx,
                             key="set_dpx_off", disabled=True, help="Applies to Duplex typologies only.")
            new_dpx = cur_dpx

        old = st.session_state.fm_params
        esc_changed = abs(float(old["escalation"].get(s_type, 0.0)) - new_e) > 1e-9
        dpx_changed = abs(float(old.get("duplex_premium", 0.0)) - new_dpx) > 1e-9

        new_esc_map = dict(params["escalation"]); new_esc_map[s_type] = new_e
        new_tr_map  = dict(params["terrace"])
        if gkey is not None and new_t is not None:
            new_tr_map[gkey] = new_t / 100.0
        np_ = {"escalation": new_esc_map, "terrace": new_tr_map,
               "duplex_premium": new_dpx, "area": dict(params.get("area", {})),
               "parking": dict(params.get("parking", {})), "base": dict(params.get("base", {}))}
        if np_ != old:
            st.session_state.fm_params = np_
            n = 0
            if esc_changed:
                n += reladder_typology(s_type, np_)                 # re-ladder this typology live
            if dpx_changed:
                for dt in ("3 Bedroom Duplex", "4 Bedroom Duplex", "5 Bedroom Duplex"):
                    n += reladder_typology(dt, np_)
            if esc_changed or dpx_changed:
                sync_floor_rates()
                st.session_state["flash"] = ("success",
                    f"✅ Escalation updated — {n} Available unit price(s) re-laddered.")
            st.rerun()

        # ── Base Price (per type; single source of truth) ─────────────────────
        st.markdown("**Base Price** — the **lowest Available** unit of this typology takes this "
                    "price/sqft; every Available unit above it follows *base + escalation × floors up*. "
                    "Sold units stay fixed.")
        base_key = PRICE_FAMILY.get(s_type, s_type)
        low_psf = lowest_available_psf(s_type, st.session_state.units)
        if low_psf is None:
            st.info(f"No {s_type} units yet — add some on a floor first, then set a base price.")
        else:
            stored_psf = get_base(s_type, params)
            default_psf = float(stored_psf) if stored_psf is not None else low_psf
            bp1, bp2 = st.columns([2, 1])
            base_psf_in = bp1.number_input(
                "Base Price (AED/sqft)", min_value=0.0, step=10.0, value=round(default_psf, 2),
                key=f"base_psf_{base_key}",
                help=f"Current lowest Available {s_type} price/sqft is AED {low_psf:,.0f}.")
            n_chg, delta = base_preview(s_type, base_psf_in, params)
            sign = "+" if delta >= 0 else "−"
            bp1.caption(f"Applying will change **{n_chg}** Available unit(s) · "
                        f"portfolio Δ **{sign}{abs(delta):,.0f} AED**.")
            if bp2.button("Apply base", key=f"apply_base_{base_key}", use_container_width=True):
                p2 = {**st.session_state.fm_params,
                      "base": {**st.session_state.fm_params.get("base", {}), base_key: base_psf_in}}
                st.session_state.fm_params = p2
                nch = recompute_from_base(s_type, base_psf_in, p2)
                sync_floor_rates()
                st.session_state["flash"] = ("success",
                    f"✅ Base price set for {s_type} — {nch} Available unit(s) recomputed.")
                st.rerun()
            if stored_psf is not None:
                st.caption(f"🔗 Active base: **AED {stored_psf:,.0f}/sqft** — escalation now "
                           "re-prices from this base.")
                if st.button("Clear base price (back to ladder anchor)", key=f"clear_base_{base_key}"):
                    p2 = {**st.session_state.fm_params}
                    p2["base"] = {k: v for k, v in p2.get("base", {}).items() if k != base_key}
                    st.session_state.fm_params = p2
                    st.session_state["flash"] = ("success", f"Base price cleared for {s_type}.")
                    st.rerun()

        st.caption("Changing **Escalation** re-prices the Available units of that typology up the "
                   "ladder immediately (Sold units stay fixed; the entry/anchor floor holds). "
                   "**Terrace %** and **Area** reprice every unit of the type. The table shows the "
                   "current settings for all typologies.")
        ref_rows = []
        for t in UNIT_TYPES:
            gk = terrace_group_key(t)
            trv = params["terrace"][gk] * 100 if gk else 100.0
            bpsf = get_base(t, params)
            base_txt = f"AED {bpsf:,.0f}" if bpsf is not None else "—"
            ref_rows.append({"Typology": t,
                             "Escalation (AED/sqft)": f"{params['escalation'].get(t, 0):,.0f}",
                             "Terrace %": f"{trv:.0f}%",
                             "Base /sqft": base_txt})
        _ref_df = pd.DataFrame(ref_rows)
        table_with_export(_ref_df, "Escalation_Terrace_Settings.xlsx", "exp_settings",
                          title="Escalation & Terrace Settings")

    with st.expander("📐  Area & Parking Settings (Internal/External sqft & parking per topology — cascades to all units of that type)", expanded=False):
        st.caption("Change a topology's area or parking and every unit of that type updates — sellable area, price, parking and all stats recompute.")
        cur_area = params.get("area", {})
        cur_park = params.get("parking", {})
        new_area, new_park = {}, {}
        ah1, ah2, ah3, ah4 = st.columns([2, 1.4, 1.4, 1.1])
        ah1.markdown("**Topology**"); ah2.markdown("**Internal (sqft)**")
        ah3.markdown("**External (sqft)**"); ah4.markdown("**Parking**")
        for t in UNIT_TYPES:
            a = cur_area.get(t, {"internal": TYPE_DEFAULTS[t]["internal"], "external": TYPE_DEFAULTS[t]["external"]})
            p = int(cur_park.get(t, TYPE_DEFAULTS[t]["parking"]))
            r1, r2, r3, r4 = st.columns([2, 1.4, 1.4, 1.1])
            r1.markdown(f"<div style='padding-top:6px'>{t}</div>", unsafe_allow_html=True)
            ni = r2.number_input("int", min_value=0.0, step=10.0, value=float(a["internal"]),
                                 key=f"area_int_{t}", label_visibility="collapsed")
            ne = r3.number_input("ext", min_value=0.0, step=10.0, value=float(a["external"]),
                                 key=f"area_ext_{t}", label_visibility="collapsed")
            np_ = r4.number_input("park", min_value=0, step=1, value=p,
                                  key=f"area_park_{t}", label_visibility="collapsed")
            new_area[t] = {"internal": ni, "external": ne}
            new_park[t] = int(np_)
        if new_area != params.get("area", {}) or new_park != params.get("parking", {}):
            st.session_state.fm_params = {**st.session_state.fm_params, "area": new_area, "parking": new_park}
            st.rerun()

    with st.expander("📈  Bulk Update by Typology & Floor Range (escalation + terrace %)", expanded=False):
        st.caption("Pick a typology and a floor range (From → To, or All floors), then optionally "
                   "**add escalation** (AED/sqft) and/or **set the terrace %**. Both apply only to "
                   "**Available** units of that typology in range — Sold units are never changed.")
        u_all = st.session_state.units
        u_all_fn = pd.to_numeric(u_all["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                                 errors="coerce")

        bt1, bt2 = st.columns([2, 1.4])
        be_type  = bt1.selectbox("Typology", UNIT_TYPES, key="be_type")
        be_scope = bt2.selectbox("Floors", ["All floors", "Floor range", "Single floor"], key="be_scope")

        # floors that actually have an Available unit of this typology
        elig = u_all[(u_all["Type"] == be_type) & (u_all["Status"] == "Available")].copy()
        elig_fn = sorted(set(pd.to_numeric(
            elig["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
            errors="coerce").dropna().astype(int)))

        # floor → "302 (2BR), 303 (2BR), 301 (3BR)" so the dropdowns show what's on each floor
        _tmp = u_all.assign(_fn=u_all_fn).dropna(subset=["_fn"]).copy()
        _tmp["_un"] = pd.to_numeric(_tmp["Unit"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                                    errors="coerce")
        _floor_units = {}
        for f, g in _tmp.groupby("_fn"):
            g = g.sort_values("_un")
            _floor_units[int(f)] = ", ".join(
                f"{r['Unit']} ({TYPE_ABBR.get(r['Type'], r['Type'])})" for _, r in g.iterrows())
        def floor_label(f):
            info = _floor_units.get(int(f))
            return f"{ordinal(f)} — {info}" if info else ordinal(f)

        if not elig_fn:
            st.info(f"No Available {be_type} units to update.")
        else:
            if be_scope == "All floors":
                f_from, f_to = elig_fn[0], elig_fn[-1]
                st.caption(f"Range: **{ordinal(f_from)} → {ordinal(f_to)}** (all {len(elig_fn)} eligible floor(s)).")
            elif be_scope == "Single floor":
                f_one = st.selectbox("Floor", elig_fn, format_func=floor_label, key="be_one")
                f_from = f_to = f_one
            else:  # Floor range
                fc1, fc2 = st.columns(2)
                f_from = fc1.selectbox("From floor", elig_fn, index=0,
                                       format_func=floor_label, key="be_from")
                to_opts = [f for f in elig_fn if f >= f_from]
                f_to = fc2.selectbox("To floor", to_opts, index=len(to_opts) - 1,
                                     format_func=floor_label, key="be_to")

            ac1, ac2, ac3 = st.columns([1.2, 1, 1.2])
            do_esc = ac1.checkbox("Add escalation", value=True, key="be_do_esc")
            amount = ac2.number_input("AED/sqft", value=100.0, step=10.0,
                                      key="be_amount", disabled=not do_esc)
            cur_tr = terrace_for(be_type, params) * 100
            tr_opts = [0, 30, 55, 65, 75, 100]
            if round(cur_tr) not in tr_opts:
                tr_opts = sorted(set(tr_opts + [int(round(cur_tr))]))
            do_tr = ac3.checkbox("Set terrace %", value=False, key="be_do_tr")
            tr_pct = ac3.selectbox("Terrace %", tr_opts,
                                   index=tr_opts.index(int(round(cur_tr))) if int(round(cur_tr)) in tr_opts else 0,
                                   key="be_trpct", disabled=not do_tr)

            mask = ((u_all["Type"] == be_type) & (u_all["Status"] == "Available") &
                    (u_all_fn >= f_from) & (u_all_fn <= f_to))
            n_hit = int(mask.sum())
            bits = []
            if do_esc: bits.append(f"escalate **{amount:+,.0f} AED/sqft**")
            if do_tr:  bits.append(f"set terrace **{tr_pct}%**")
            action_txt = " and ".join(bits) if bits else "make no change (tick an action)"
            st.caption(f"Will update **{n_hit}** Available {be_type} unit(s) on floors "
                       f"{ordinal(f_from)}–{ordinal(f_to)}: {action_txt}.")

            can_apply = n_hit > 0 and (do_esc or do_tr)
            if st.button("Apply", type="primary", disabled=not can_apply, key="be_apply"):
                if do_esc:
                    st.session_state.units.loc[mask, "Price_sqft"] = \
                        st.session_state.units.loc[mask, "Price_sqft"] + amount
                if do_tr:
                    st.session_state.units.loc[mask, "Terrace_Override"] = tr_pct / 100.0
                # keep floor-object rates in sync for the floor table / export
                for fl in st.session_state.floors:
                    if f_from <= fl["floor"] <= f_to:
                        for un in fl["units"]:
                            uid = un.get("uid")
                            if uid is not None:
                                r = st.session_state.units[st.session_state.units["uid"] == uid]
                                if not r.empty and r.iloc[0]["Type"] == be_type and r.iloc[0]["Status"] == "Available":
                                    un["rate"] = float(r.iloc[0]["Price_sqft"])
                st.session_state["flash"] = ("success",
                    f"✅ Updated {n_hit} Available {be_type} unit(s) on floors "
                    f"{ordinal(f_from)}–{ordinal(f_to)}: {action_txt.replace('**','')}.")
                st.rerun()

    # Floor totals (the Floors table was removed; grand is still needed for previews below)
    grand = sum(floor_total(fl, params) for fl in floors)
    res_floor_nums = {fl["floor"] for fl in floors}
    n_res = len(floors)
    n_mep = len([f for f in blocked if f not in res_floor_nums])     # MEP/Majlis floors
    mc1, mc2 = st.columns(2)
    mc1.metric("Floors", n_res + n_mep)
    mc2.metric("Units",  sum(len(fl["units"]) for fl in floors))
    mc1.caption(f"{n_res} residential + {n_mep} MEP/Majlis")

    st.divider()
    action = st.radio("Action",
                      ["Add a New Floor", "Insert a Floor (between)", "Remove Floor(s)", "Edit a Floor"],
                      horizontal=True, key="fm_action")
    st.divider()

    # ─────────────────────── ADD A NEW FLOOR ──────────────────────────────────
    if action == "Add a New Floor":
        st.subheader("Add New Floor(s)")
        st.caption("Pick a floor range (set **From = To** for a single floor) and one unit mix — the "
                   "**same mix** is added to every floor in the range. Units are priced from the escalation "
                   "ladder for their floor (built bottom-up). Blocked / existing floors are skipped.")
        st.caption("🔼 A **Duplex** is one unit that occupies the floor you add it on **plus the floor above** "
                   "(the roof shifts up automatically) — so add a duplex on a **single floor**; it counts as "
                   "**1 unit** even though it shows as two levels in the Building View.")
        existing = [fl["floor"] for fl in floors]
        # an ADDED (upward) duplex physically occupies its floor + the one ABOVE, so that upper
        # level is taken too and must not be offered as a new floor
        u_all = st.session_state.units
        dup_up_tops = set()
        if "Dup_Up" in u_all.columns:
            u_fn = pd.to_numeric(u_all["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                                 errors="coerce")
            upmask = (u_all["Type"].astype(str).str.contains("Duplex") &
                      u_all["Dup_Up"].apply(lambda v: pd.notna(v) and bool(v)))
            dup_up_tops = set((u_fn[upmask].dropna().astype(int) + 1).tolist())
        occupied = set(existing) | dup_up_tops
        ac1, ac2 = st.columns(2)
        nf_from = ac1.number_input("From floor", min_value=1, max_value=999,
                                   value=(max(occupied)+1 if occupied else 59), step=1, key="newfl_from")
        nf_to = ac2.number_input("To floor", min_value=1, max_value=999,
                                 value=int(nf_from), step=1, key="newfl_to")
        lo, hi = int(min(nf_from, nf_to)), int(max(nf_from, nf_to))

        st.markdown("**Unit mix** — pick topology and use **− / +** to set quantity (applied to each floor):")
        mix = unit_mix_builder("addmix", [{"type": "3 Bedroom", "qty": 1}, {"type": "2 Bedroom", "qty": 2}])
        mix = [(t, (1 if "Duplex" in t else q)) for t, q in mix]      # a duplex is always 1 unit per floor
        has_duplex = any(("Duplex" in t) and q > 0 for t, q in mix)

        if has_duplex:
            # a duplex ALWAYS occupies From + (From+1) as a single unit — the "To floor" is ignored
            base, upper = lo, lo + 1
            st.info(f"🔼 Duplex selected → it will occupy **{ordinal(base)}–{ordinal(upper)}** as **one** unit "
                    f"(the *To floor* is ignored; a duplex is always exactly two levels, max one per floor).")
            conflict = [f for f in (base, upper) if f in blocked or f in occupied]
            if conflict:
                st.error("Cannot place the duplex — these level(s) are blocked or already occupied: "
                         + ", ".join(ordinal(f) for f in conflict))
                place_floors = []
            else:
                place_floors = [base]                    # one record floor; the upper level is auto-reserved
            total_units = sum(q for _, q in mix)
            approx = sum(unit_val(t, new_unit_rate(t, base, st.session_state.units, params), params)["total"] * q
                         for t, q in mix)
            m1, m2, m3 = st.columns(3)
            m1.metric("Levels occupied", f"{base}–{upper}" if place_floors else "—")
            m2.metric("Units to add", total_units)
            m3.metric("Added value (≈)", aed(approx))
            add_label = f"Add Duplex on {ordinal(base)} (occupies {ordinal(base)}–{ordinal(upper)})"
        else:
            rng = list(range(lo, hi + 1))
            place_floors = [f for f in rng if f not in blocked and f not in occupied]
            skip_block   = [f for f in rng if f in blocked]
            skip_exists  = [f for f in rng if f in occupied]
            if skip_block:
                st.warning("Skipping blocked (MEP/Majlis): " + ", ".join(ordinal(f) for f in skip_block))
            if skip_exists:
                st.warning("Skipping floors already occupied (incl. the upper level of a duplex): "
                           + ", ".join(ordinal(f) for f in skip_exists))
            if not place_floors:
                st.error("No valid floors in this range (all are blocked or already occupied).")
            per_floor = sum(q for _, q in mix)
            total_units = per_floor * len(place_floors)
            approx = 0.0
            for f in place_floors:
                for t, q in mix:
                    approx += unit_val(t, new_unit_rate(t, f, st.session_state.units, params), params)["total"] * q
            m1, m2, m3 = st.columns(3)
            m1.metric("Floors to add", len(place_floors))
            m2.metric("Units to add", total_units)
            m3.metric("Added value (≈)", aed(approx))
            if place_floors:
                st.caption(f"Floors to add: {', '.join(ordinal(f) for f in place_floors)}")
            add_label = ((f"Add Floor {ordinal(place_floors[0])}" if len(place_floors) == 1
                          else f"Add {len(place_floors)} floors "
                               f"({ordinal(min(place_floors))}–{ordinal(max(place_floors))})")
                         if place_floors else "Add")

        can_add = len(place_floors) > 0 and sum(q for _, q in mix) > 0
        if st.button(add_label, type="primary", key="btn_addfl", disabled=not can_add):
            try:
                ordered = []
                for t, q in mix:
                    ordered += [t] * q
                ordered.sort(key=lambda t: (t != "3 Bedroom", t))
                for f in place_floors:                    # ascending → each floor escalates off the one below
                    nos = gen_unit_nos(f, ordered)
                    new_units = [{"unit_no": no, "type": t,
                                  "rate": new_unit_rate(t, f, st.session_state.units, params)}
                                 for no, t in zip(nos, ordered)]
                    add_units_to_register(new_units, f, params)
                    st.session_state.floors.append({"floor": f, "kind": "Added",
                                                    "levels": max(TYPE_DEFAULTS[t]["levels"] for t in ordered),
                                                    "units": new_units})
                st.session_state.floors.sort(key=lambda x: x["floor"])
                st.session_state.pop("_mep_normalized", None)   # re-clean the whole tower next run
                clear_builder("addmix")
                for _k in ("newfl_from", "newfl_to"):
                    st.session_state.pop(_k, None)
                st.session_state["flash"] = ("success",
                    f"✅ Added {len(place_floors)} floor record(s), {total_units} unit(s).")
            except Exception as e:
                st.session_state["flash"] = ("error", f"❌ Could not add floors: {e}")
            st.rerun()

    # ─────────────────────── INSERT FLOOR(S) (BETWEEN) ─────────────────────────
    elif action == "Insert a Floor (between)":
        st.subheader("Insert Floor(s) in between")
        st.caption("Insert **one or more** brand-new floors starting at a chosen level. Those floors and "
                   "**everything above them move up** — floor numbers **and** unit numbers are renumbered — "
                   "while **MEP / Majlis floors keep their fixed numbers** (residential renumbers around "
                   "them). The roof rides up; moved units keep their prices; new floors are priced from "
                   "the escalation ladder.")
        res_floors = sorted({fl["floor"] for fl in floors
                             if fl["floor"] not in blocked and fl["floor"] != 2})
        if not res_floors:
            st.info("No residential floors to insert between.")
        else:
            ic1, ic2 = st.columns(2)
            ins_from = ic1.selectbox("Insert starting at level", res_floors,
                                     format_func=lambda f: f"{ordinal(f)}", key="ins_from")
            n_floors = ic2.number_input("How many floors to insert", min_value=1, max_value=50,
                                        value=1, step=1, key="ins_count")
            st.markdown("**Unit mix for each new floor** — the same mix is applied to every inserted floor:")
            imix = unit_mix_builder("insmix", [{"type": "2 Bedroom", "qty": 2}, {"type": "3 Bedroom", "qty": 1}])
            has_dup = any(("Duplex" in t) and q > 0 for t, q in imix)
            per = sum(q for _, q in imix)
            # the actual new floor numbers (skipping any MEP/Majlis levels in the way)
            _fixed = set(blocked) | {2}
            _slots, _x = [], int(ins_from)
            while len(_slots) < int(n_floors):
                while _x in _fixed:
                    _x += 1
                _slots.append(_x); _x += 1
            if has_dup:
                st.error("Duplex units can't be inserted in between (a duplex occupies two levels). "
                         "Add duplexes from **Add a New Floor** instead.")
            m1, m2, m3 = st.columns(3)
            m1.metric("New floors", int(n_floors))
            m2.metric("Units to add", per * int(n_floors))
            m3.metric("New floor levels",
                      ", ".join(ordinal(s) for s in _slots[:5]) + ("…" if len(_slots) > 5 else ""))
            n_above = len([f for f in res_floors if f >= int(ins_from)])
            st.caption(f"⚠️ This renumbers **{n_above}** residential floor(s) at/above {ordinal(ins_from)} "
                       f"(+{int(n_floors)} level(s), skipping MEP/Majlis). MEP / Majlis floors are unchanged.")
            can_ins = (per > 0) and (not has_dup)
            if st.button(f"Insert {int(n_floors)} floor(s) at {ordinal(ins_from)}", type="primary",
                         key="btn_insfl", disabled=not can_ins):
                try:
                    ordered = []
                    for t, q in imix:
                        ordered += [t] * q
                    ordered.sort(key=lambda t: (t != "3 Bedroom", t))
                    new_nums, remap = insert_floors_between(int(ins_from), int(n_floors), ordered, params)
                    st.session_state.pop("_mep_normalized", None)   # re-clean the whole tower next run
                    clear_builder("insmix")
                    for _k in ("ins_from", "ins_count"):
                        st.session_state.pop(_k, None)
                    st.session_state["flash"] = ("success",
                        f"✅ Inserted {len(new_nums)} floor(s) at {', '.join(ordinal(n) for n in new_nums)}; "
                        f"{len(remap)} floor(s) above renumbered. MEP / Majlis floors unchanged.")
                except Exception as e:
                    st.session_state["flash"] = ("error", f"❌ Could not insert: {e}")
                st.rerun()

    # ─────────────────────── REMOVE FLOOR(S) (BETWEEN) ─────────────────────────
    elif action == "Remove Floor(s)":
        st.subheader("Remove Floor(s) in between")
        st.caption("Remove a **range** of floors. Every residential floor **above** the range cascades "
                   "**down** to fill the gap — floor numbers **and** unit numbers renumber — while "
                   "**MEP / Majlis floors stay fixed**. If **any** unit in the range is **Sold**, removal "
                   "is blocked.")
        res_floors = sorted({fl["floor"] for fl in floors
                             if fl["floor"] not in blocked and fl["floor"] != 2})
        if not res_floors:
            st.info("No residential floors to remove.")
        else:
            u_all = st.session_state.units
            u_fn = pd.to_numeric(u_all["Floor"].astype(str).str.replace(r"[^0-9]", "", regex=True),
                                 errors="coerce")
            floor_units = {}
            for _i in u_all.index:
                _f = u_fn[_i]
                if pd.notna(_f):
                    floor_units.setdefault(int(_f), []).append(str(u_all.at[_i, "Unit"]))
            def _fmt_floor(f):
                us = floor_units.get(f, [])
                return f"{ordinal(f)}  ·  {', '.join(us)}" if us else ordinal(f)

            rc1, rc2 = st.columns(2)
            rm_from = rc1.selectbox("From floor", res_floors, format_func=_fmt_floor, key="rm_from")
            to_opts = [f for f in res_floors if f >= rm_from]
            rm_to = rc2.selectbox("To floor", to_opts, index=0, format_func=_fmt_floor, key="rm_to")
            lo, hi = int(rm_from), int(rm_to)
            in_range = [f for f in res_floors if lo <= f <= hi]

            rng_units = u_all[u_fn.isin(in_range)]
            sold = rng_units[rng_units["Status"] == "Sold"]
            n_above = len([f for f in res_floors if f > hi])

            m1, m2, m3 = st.columns(3)
            m1.metric("Floors to remove", len(in_range))
            m2.metric("Units to remove", len(rng_units))
            m3.metric("Floors cascading down", n_above)
            st.caption(f"Floors to remove: {', '.join(ordinal(f) for f in in_range)}  ·  "
                       f"{n_above} floor(s) above will drop down (MEP/Majlis unchanged).")

            if not sold.empty:
                _sl = ", ".join(f"{r['Unit']} ({r['Floor']})" for _, r in sold.iterrows())
                st.error(f"🚫 Removal blocked — **{len(sold)} Sold unit(s)** in this range: {_sl}. "
                         "Sold units can't be removed; re-mark them Available or choose another range.")
            else:
                st.warning(f"This permanently removes {len(rng_units)} unit(s) on {len(in_range)} "
                           f"floor(s) and renumbers everything above. Use Reload/Base Version to restore.")

            can_rm = (len(in_range) > 0) and sold.empty
            if st.button(f"Remove {len(in_range)} floor(s) ({ordinal(lo)}–{ordinal(hi)})",
                         type="primary", key="btn_rmfl", disabled=not can_rm):
                try:
                    removed, remap = remove_floors_between(lo, hi)
                    st.session_state.pop("_mep_normalized", None)   # re-clean the whole tower next run
                    for _k in ("rm_from", "rm_to"):
                        st.session_state.pop(_k, None)
                    st.session_state["flash"] = ("success",
                        f"✅ Removed {len(removed)} floor(s) ({ordinal(lo)}–{ordinal(hi)}); "
                        f"{len(remap)} floor(s) above cascaded down. MEP / Majlis unchanged.")
                except Exception as e:
                    st.session_state["flash"] = ("error", f"❌ Could not remove: {e}")
                st.rerun()

    # ─────────────────────── EDIT A FLOOR ─────────────────────────────────────
    else:
        st.subheader("Edit Floor(s)")
        st.caption("Pick a floor — or a **From → To** range — then set the **Available** unit mix "
                   "(add, remove or swap types). Sold units are always protected. In range mode the "
                   "same Available mix is applied to every floor in range that has available units.")
        floor_nums = [fl["floor"] for fl in floors]
        _fl_labels = {
            fl["floor"]: f"Floor {ordinal(fl['floor'])}  —  " + ", ".join(
                f"{u['unit_no']} ({TYPE_ABBR.get(u['type'], u['type'])})" for u in fl["units"])
            for fl in floors
        }
        ec1, ec2 = st.columns(2)
        sel = ec1.selectbox(
            "From floor", ["— select —"] + floor_nums,
            format_func=lambda x: _fl_labels.get(x, f"Floor {ordinal(x)}") if x != "— select —" else x,
            key="edit_floor_sel",
        )
        sel_to = sel
        if sel != "— select —":
            to_opts = [f for f in floor_nums if f >= sel]
            sel_to = ec2.selectbox("To floor", to_opts, index=0,
                                   format_func=lambda x: _fl_labels.get(x, f"Floor {ordinal(x)}"),
                                   key="edit_floor_to")

        if sel != "— select —":
            lo, hi = int(sel), int(sel_to)
            range_floors = [f for f in floor_nums if lo <= f <= hi]
            is_range = len(range_floors) > 1
            smap = uid_status_map()

            def apply_mix_to_floor(floor_num, mix):
                """Replace a floor's Available units with `mix` (keep Sold). Returns (added, removed, locked, skipped)."""
                fobj = next((f for f in st.session_state.floors if f["floor"] == floor_num), None)
                if fobj is None:
                    return (0, 0, 0, True)
                locked, avail = split_floor_units(fobj)
                if not avail:
                    return (0, 0, 0, True)                       # fully sold → skip
                keep, new_meta = [], []
                for t, q in mix:
                    same = [u for u in avail if u["type"] == t]
                    for j in range(q):
                        (keep.append(same[j]) if j < len(same) else new_meta.append(t))
                kept = {u["uid"] for u in keep}
                removed = [u["uid"] for u in avail if u["uid"] not in kept]
                if removed:
                    remove_units_from_register(removed)
                added = []
                if new_meta:
                    new_meta.sort(key=lambda t: (t != "3 Bedroom", t))
                    nos = gen_unit_nos(floor_num, new_meta)
                    added = [{"unit_no": no, "type": t,
                              "rate": new_unit_rate(t, floor_num, st.session_state.units, params)}
                             for no, t in zip(nos, new_meta)]
                    add_units_to_register(added, floor_num, params)
                final = locked + keep + added
                if final:
                    fobj["units"] = final
                    if fobj["kind"] not in ("Pool",):
                        fobj["kind"] = "Edited"
                else:
                    st.session_state.floors = [f for f in st.session_state.floors if f["floor"] != floor_num]
                return (len(added), len(removed), len(locked), False)

            def floor_mix_editor(floor_num, key_suffix):
                """Show a floor's current units + an editable Available mix builder.
                Returns (mix, has_available)."""
                fobj = next(x for x in floors if x["floor"] == floor_num)
                lk, av = split_floor_units(fobj)
                cur = [{"Unit": u["unit_no"], "Type": u["type"], "Status": unit_status(u, smap),
                        "Editable": "Yes" if unit_status(u, smap) == "Available" else "🔒 No",
                        "Rate/sqft": f"AED {u['rate']:,.0f}",
                        "Value": aed(unit_val(u["type"], u["rate"], params)["total"])} for u in fobj["units"]]
                excel_table(pd.DataFrame(cur))
                if not av:
                    st.caption("🔒 No Available units on this floor — it will be skipped.")
                    return [], False
                if lk:
                    st.caption(f"{len(lk)} Sold unit(s) kept; editing the {len(av)} Available unit(s).")
                comp = {}
                for u in av:
                    comp[u["type"]] = comp.get(u["type"], 0) + 1
                drows = [{"type": t, "qty": q} for t, q in comp.items()] or [{"type": "2 Bedroom", "qty": 1}]
                return unit_mix_builder(f"editmix_{key_suffix}", drows, qmin=0), True

            if is_range:
                aff = [f for f in range_floors
                       if any(unit_status(u, smap) == "Available"
                              for u in next(x for x in floors if x["floor"] == f)["units"])]
                st.markdown(f"**Editing floors {ordinal(lo)}–{ordinal(hi)} together** — set **one** "
                            f"Available unit mix below and it's applied to **every floor in range** that "
                            f"has available units ({len(aff)} floor(s)). Sold units on each floor are kept.")
                st.markdown("**Available unit mix — use − / + to set quantity (applied to each floor):**")
                mix = unit_mix_builder(f"editmix_rng_{lo}_{hi}", [{"type": "2 Bedroom", "qty": 1}], qmin=0)
                if not aff:
                    st.error("No floors with available units in this range.")
                elif st.button(f"Apply to {len(aff)} floor(s)", type="primary", key="btn_edit_apply_rng"):
                    try:
                        tot_add = tot_rem = 0
                        for f in aff:                          # ascending so escalation builds up
                            a, r, l, sk = apply_mix_to_floor(f, mix)
                            tot_add += a; tot_rem += r
                        st.session_state.floors.sort(key=lambda x: x["floor"])
                        clear_builder(f"editmix_rng_{lo}_{hi}")
                        st.session_state["flash"] = ("success",
                            f"✅ Updated {len(aff)} floor(s) {ordinal(lo)}–{ordinal(hi)} — "
                            f"{tot_add} added, {tot_rem} removed.")
                    except Exception as e:
                        st.session_state["flash"] = ("error", f"❌ Could not update range: {e}")
                    st.rerun()
            else:
                fl = next(f for f in floors if f["floor"] == sel)
                locked_units, avail_units = split_floor_units(fl)
                st.markdown(f"**Floor {ordinal(sel)}** ({fl['kind']}, {fl['levels']} level"
                            f"{'s' if fl['levels']>1 else ''}) — current units:")
                mix, has_av = floor_mix_editor(sel, f"{sel}")

                if not has_av:
                    st.error(f"🔒 Floor {ordinal(sel)} has no Available units — pick a floor with available units.")
                else:
                    locked_total = sum(unit_val(u["type"], u["rate"], params)["total"] for u in locked_units)
                    avail_total = 0
                    for t, q in mix:
                        same = [u for u in avail_units if u["type"] == t]
                        for j in range(q):
                            if j < len(same):
                                avail_total += unit_val(t, same[j]["rate"], params)["total"]
                            else:
                                avail_total += unit_val(t, new_unit_rate(t, sel, st.session_state.units, params), params)["total"]
                    new_total = locked_total + avail_total
                    old_total = floor_total(fl, params)
                    grand_excl = grand - old_total

                    m1, m2, m3 = st.columns(3)
                    m1.metric("Units after edit", len(locked_units) + sum(q for _, q in mix))
                    m2.metric("Floor value", aed(new_total), delta=aed(new_total - old_total))
                    m3.metric("New Portfolio", aed(grand_excl + new_total))

                    b1, b2 = st.columns(2)
                    if b1.button("Apply Changes", type="primary", key="btn_edit_apply"):
                        try:
                            a, r, l, sk = apply_mix_to_floor(sel, mix)
                            clear_builder(f"editmix_{sel}")
                            st.session_state["flash"] = ("success",
                                f"✅ Floor {ordinal(sel)} updated — {a} added, {r} removed, "
                                f"{l} protected unit(s) kept.")
                        except Exception as e:
                            st.session_state["flash"] = ("error", f"❌ Could not update floor {sel}: {e}")
                        st.rerun()

                    if locked_units:
                        b2.button("Remove Entire Floor", key="btn_edit_remove", disabled=True,
                                  help="Cannot remove a floor that has Sold units.")
                    elif b2.button("Remove Entire Floor", key="btn_edit_remove"):
                        remove_units_from_register([u["uid"] for u in fl["units"]])
                        st.session_state.floors = [f2 for f2 in st.session_state.floors if f2["floor"] != sel]
                        st.session_state.pop("_mep_normalized", None)   # re-clean the whole tower next run
                        clear_builder(f"editmix_{sel}")
                        st.session_state["flash"] = ("success", f"✅ Floor {ordinal(sel)} removed entirely.")
                        st.rerun()


# ── Tab 4: Edit / Remove individual units ──────────────────────────────────────

with tab4:
    def uid_label(uid, frame):
        r = frame[frame["uid"] == uid]
        if r.empty:
            return uid
        r = r.iloc[0]
        return f"Unit {r['Unit']} - {r['Type']} (Floor {r['Floor']}) - {r['Status']}"

    st.subheader("Edit Unit(s)")
    st.caption("Pick a **unit range** — From → To (set To = From to edit a single unit). The chosen "
               "changes apply to **every unit in the range**. Saving stays on this page — navigate "
               "wherever you like afterwards. All units are editable, including Sold.")

    uids = df["uid"].tolist()
    eu1, eu2 = st.columns(2)
    from_uid = eu1.selectbox("From unit", ["— select —"] + uids,
                             format_func=lambda x: uid_label(x, df) if x != "— select —" else x,
                             key="edit_from")
    range_uids = []
    if from_uid != "— select —":
        fi = uids.index(from_uid)
        to_opts = uids[fi:]                                    # To must be at/after From
        to_uid = eu2.selectbox("To unit", to_opts, index=0,
                               format_func=lambda x: uid_label(x, df), key="edit_to")
        range_uids = uids[fi:uids.index(to_uid) + 1]

    if range_uids:
        n_sel = len(range_uids)
        st.caption(f"**{n_sel} unit(s) selected** — {uid_label(range_uids[0], df)}"
                   + (f"  →  {uid_label(range_uids[-1], df)}" if n_sel > 1 else ""))
        st.caption("Edit **each unit individually** below. Entering a **Total Value** instantly "
                   "updates that row's **Price / sellable sqft** in the table (and editing Price or "
                   "Sellable updates Total) — but nothing is written until you press **Save Changes**. "
                   "An edited Sellable sqft is kept as a per-unit override.")

        # `base` = current (saved) snapshot — used to detect what actually changed on Save
        rng = df[df["uid"].isin(range_uids)].copy()
        rng["uid"] = pd.Categorical(rng["uid"], categories=range_uids, ordered=True)
        rng = rng.sort_values("uid")
        base = pd.DataFrame({
            "Unit":  rng["Unit"].astype(str).values,
            "Type":  rng["Type"].astype(str).values,
            "Floor": rng["Floor"].astype(str).values,
            "Status": rng["Status"].astype(str).values,
            "Price/sellable sqft": rng["Price_sqft"].round(0).astype(float).values,
            "Sellable sqft": rng["Sellable_sqft"].round(0).astype(float).values,
            "Total Value": rng["Price"].round(0).astype(float).values,
            "_uid":  rng["uid"].astype(str).values,
        })
        rngkey   = f"{range_uids[0]}_{range_uids[-1]}"
        work_key = f"ed_work_{rngkey}"          # live preview copy (not yet persisted)
        ekey     = f"ed_editor_{rngkey}"
        if work_key not in st.session_state:
            st.session_state[work_key] = base.copy()

        def _recompute_edit():
            wdf = st.session_state[work_key]
            delta = st.session_state[ekey].get("edited_rows", {})
            for ridx, chg in delta.items():
                ridx = int(ridx)
                for col, val in chg.items():
                    wdf.at[ridx, col] = val
                sell = float(wdf.at[ridx, "Sellable sqft"] or 0)
                if "Total Value" in chg:                       # Total entered → recompute price/sellable
                    wdf.at[ridx, "Price/sellable sqft"] = float(round(wdf.at[ridx, "Total Value"] / sell)) if sell else 0.0
                elif "Sellable sqft" in chg or "Price/sellable sqft" in chg:   # either → recompute Total
                    wdf.at[ridx, "Total Value"] = float(round(wdf.at[ridx, "Price/sellable sqft"] * sell))
            st.session_state[work_key] = wdf.copy()            # new identity → editor re-reads recomputed values
            st.session_state[ekey]["edited_rows"] = {}         # consume delta so the table shows recomputed values

        st.data_editor(
            st.session_state[work_key], hide_index=True, use_container_width=True,
            key=ekey, on_change=_recompute_edit,
            column_config={
                "Unit":  st.column_config.TextColumn("Unit", disabled=True),
                "Type":  st.column_config.TextColumn("Type", disabled=True),
                "Floor": st.column_config.TextColumn("Floor", disabled=True),
                "Status": st.column_config.SelectboxColumn("Status", options=STATUS_OPTIONS, required=True),
                "Price/sellable sqft": st.column_config.NumberColumn(
                    "Price / sellable sqft (AED)", min_value=0.0, step=50.0, format="%.0f"),
                "Sellable sqft": st.column_config.NumberColumn(
                    "Sellable sqft", min_value=0.0, step=10.0, format="%.0f"),
                "Total Value": st.column_config.NumberColumn(
                    "Total Value (AED)", min_value=0.0, step=10000.0, format="%.0f"),
                "_uid": None,                              # hidden key column
            },
        )

        if st.button("Save Changes", type="primary", key="btn_save"):
            wdf = st.session_state[work_key]
            u = st.session_state.units
            if "Sellable_Override" not in u.columns:        # safety for older saved state
                u["Sellable_Override"] = pd.NA
            u_uid_str = u["uid"].astype(str)
            orig_sell = dict(zip(base["_uid"], base["Sellable sqft"]))
            done = 0
            for idx in wdf.index:
                uidk = str(wdf.at[idx, "_uid"])
                ix = u.index[u_uid_str == uidk]
                if len(ix) == 0:
                    continue
                i = ix[0]
                u.at[i, "Status"] = wdf.at[idx, "Status"]
                # sellable: override only when the user actually changed it (else stay dynamic)
                new_sell = float(wdf.at[idx, "Sellable sqft"])
                if abs(new_sell - float(orig_sell.get(uidk, new_sell))) > 0.5:
                    u.at[i, "Sellable_Override"] = new_sell
                # price/sellable sqft already reflects any Total Value edit (live recompute above)
                u.at[i, "Price_sqft"] = float(wdf.at[idx, "Price/sellable sqft"])
                done += 1
            st.session_state.pop(work_key, None)            # refresh preview from saved values next render
            # No st.rerun() — stay on this page; the user navigates when ready.
            st.success(f"✅ Saved per-unit changes to {done} unit(s). "
                       "Use the tabs above to navigate when ready.")

    st.divider()
    st.subheader("Appreciation / Discount")
    st.caption("Select the units to work with, then set an **appreciation** (price up) or **discount** "
               "(price down) **for each unit individually** in the grid — by **percentage** or a **total "
               "AED amount** per row. Set a row to **None** to clear that unit. Apply updates the price in "
               "the Register, Building View and portfolio totals (Floor-Wise Variance keeps the list price).")
    _adj_pool = df[df["Status"] == "Available"]["uid"].tolist()   # Sold units excluded
    adj_uids = st.multiselect("Units to adjust (available units only)", _adj_pool,
                              format_func=lambda u: uid_label(u, df), key="adj_units")
    if adj_uids:
        st.caption("Set each unit's adjustment below, then press **Apply** — every row submits together. "
                   "Value = a percentage (e.g. 5 = 5%) or a total AED amount, per the **By** column; "
                   "set **None** to clear a unit.")
        with st.form("adj_form", clear_on_submit=False):
            _hdr = st.columns([2.6, 1.4, 1.5, 1.2])
            _hdr[0].markdown("**Unit**"); _hdr[1].markdown("**Adjustment**")
            _hdr[2].markdown("**By**");   _hdr[3].markdown("**Value**")
            _rows = []
            for _uid in adj_uids:
                _r = df[df["uid"] == _uid].iloc[0]
                _cur = pd.to_numeric(_r.get("Adj_Pct"), errors="coerce")
                _di = 0 if (pd.isna(_cur) or float(_cur) == 0) else (1 if float(_cur) > 0 else 2)
                _c = st.columns([2.6, 1.4, 1.5, 1.2])
                _c[0].markdown(
                    f"**{_r['Unit']}** · {_r['Type']} · {_r['Floor']}  \n"
                    f"<span style='color:#8a8a8a;font-size:0.85em'>List AED {float(_r['Base_Price']):,.0f}</span>",
                    unsafe_allow_html=True)
                _adj = _c[1].selectbox("Adjustment", ["None", "Appreciation", "Discount"], index=_di,
                                       key=f"adj_dir_{_uid}", label_visibility="collapsed")
                _basis = _c[2].selectbox("By", ["Percentage", "Amount (AED)"],
                                         key=f"adj_basis_{_uid}", label_visibility="collapsed")
                _val = _c[3].number_input("Value", min_value=0.0, step=1.0,
                                          value=(abs(float(_cur)) if pd.notna(_cur) and float(_cur) != 0 else 0.0),
                                          key=f"adj_val_{_uid}", label_visibility="collapsed")
                _rows.append((_uid, _adj, _basis, _val))
            _b1, _b2 = st.columns(2)
            _do_apply = _b1.form_submit_button(f"Apply to {len(adj_uids)} unit(s)", type="primary",
                                               use_container_width=True)
            _do_clear = _b2.form_submit_button("Clear adjustment for these units",
                                               use_container_width=True)
        if _do_apply:
            u = st.session_state.units
            if "Adj_Pct" not in u.columns:
                u["Adj_Pct"] = pd.NA
            n_set = 0
            for _uid, _adj, _basis, _val in _rows:
                _ix = u.index[u["uid"] == _uid]
                if not len(_ix):
                    continue
                _val = float(_val or 0)
                if _adj == "None" or _val <= 0:
                    u.at[_ix[0], "Adj_Pct"] = pd.NA
                    continue
                sign = 1.0 if _adj == "Appreciation" else -1.0
                if _basis == "Percentage":
                    _pct = sign * _val
                else:                                      # total AED → % of that unit's LIST price
                    _bp = float(df.loc[df["uid"] == _uid, "Base_Price"].iloc[0])
                    _pct = sign * (_val / _bp * 100.0) if _bp else 0.0
                u.at[_ix[0], "Adj_Pct"] = round(_pct, 6); n_set += 1
            st.session_state["flash"] = ("success",
                f"Applied per-unit adjustments — {n_set} set / {len(adj_uids) - n_set} cleared.")
            st.rerun()
        if _do_clear:
            u = st.session_state.units
            if "Adj_Pct" in u.columns:
                u.loc[u["uid"].isin(adj_uids), "Adj_Pct"] = pd.NA
            st.session_state["flash"] = ("success", f"Cleared adjustment for {len(adj_uids)} unit(s).")
            st.rerun()

    _adj_now = (df[pd.to_numeric(df["Adj_Pct"], errors="coerce").fillna(0) != 0].copy()
                if "Adj_Pct" in df.columns else df.iloc[0:0])
    if len(_adj_now):
        _p = pd.to_numeric(_adj_now["Adj_Pct"], errors="coerce")
        _cur = pd.DataFrame({
            "Unit": _adj_now["Unit"].values, "Type": _adj_now["Type"].values,
            "Floor": _adj_now["Floor"].values,
            "List Price (AED)": _adj_now["Base_Price"].map(lambda v: f"AED {v:,.0f}").values,
            "Appreciation / Discount": _p.map(lambda v: ("+" if v > 0 else "−") + f"{abs(v):.1f}%").values,
            "Updated Price (AED)": _adj_now["Price"].map(lambda v: f"AED {v:,.0f}").values,
        })
        st.caption(f"**{len(_cur)} unit(s) currently adjusted** (green = appreciation, red = discount):")
        def _c_adj(val):
            s = str(val)
            if s.startswith("+"):                          return "color:#1a7f37;font-weight:600"
            if s.startswith("−") or s.startswith("-"):     return "color:#d1242f;font-weight:600"
            return ""
        st.dataframe(_cur.style.map(_c_adj, subset=["Appreciation / Discount"]),
                     hide_index=True, use_container_width=True)
    else:
        st.caption("No units currently have an appreciation / discount applied.")

    st.divider()
    st.subheader("Remove Units")
    rt = st.multiselect("Filter by type",   sorted(df["Type"].unique().tolist()), key="rem_type")
    rs = st.multiselect("Filter by status", STATUS_OPTIONS, key="rem_status")
    rdf = df.copy()
    if rt: rdf = rdf[rdf["Type"].isin(rt)]
    if rs: rdf = rdf[rdf["Status"].isin(rs)]
    torem = st.multiselect("Select unit(s) to remove", rdf["uid"].tolist(),
                           format_func=lambda u: uid_label(u, rdf), key="rem_ms")
    if torem:
        st.warning(f"This will remove {len(torem)} unit(s). Use Reload to restore.")
        if st.button("Remove Selected Units", type="primary", key="btn_rem"):
            remove_units_from_register(torem)
            st.success(f"Removed {len(torem)} unit(s).")
            st.rerun()


# ── Export helpers (per-table export buttons live on each tab) ──────────────────

def _style_export_sheet(ws):
    """Apply clean, professional formatting to a worksheet: bold blue header, frozen
    header row, AED / area / % number formats, borders, banded rows, auto widths."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin   = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    band_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")

    headers = [c.value for c in ws[1]]
    for c in ws[1]:
        c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    def fmt_for(h):
        h = str(h)
        if "AED" in h:                         return '"AED" #,##0'
        if "Terrace Rate" in h:                return "0%"
        if "sqft" in h:                        return "#,##0.0"
        return None

    for ci, h in enumerate(headers, start=1):
        fmt = fmt_for(h)
        maxlen = len(str(h))
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci)
            cell.border = border
            cell.alignment = left if ci == 1 else center
            if fmt:
                cell.number_format = fmt
            if r % 2 == 0:
                cell.fill = band_fill
            if cell.value is not None:
                maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 3, 12), 42)


def build_export():
    d = recalc(st.session_state.units, st.session_state.fm_params)
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        eu = d[["Type","Status","Unit","Floor","Parking","Internal_sqft","External_sqft",
                "Total_sqft","Sellable_sqft","Terrace_Rate","Price_sqft","Price"]].copy()
        eu.columns = ["Type","Status","Unit","Floor","Parking","Internal (sqft)","External (sqft)",
                      "Total (sqft)","Sellable (sqft)","Terrace Rate","Price/sqft (AED)","Price (AED)"]
        eu.to_excel(writer, index=False, sheet_name="Unit Register")

        grp = d.groupby("Type").agg(
            Units=("Unit","count"), Sold=("Status", lambda x: (x=="Sold").sum()),
            Available=("Status", lambda x: (x=="Available").sum()),
            Avg_Price_sqft=("Price_sqft","mean"), Min_Price_sqft=("Price_sqft","min"),
            Max_Price_sqft=("Price_sqft","max"), Total_Sellable=("Sellable_sqft","sum"),
            Total_Value=("Price","sum")).reset_index()
        grp.columns = ["Type","Units","Sold","Available","Avg Price/sqft (AED)","Min Price/sqft (AED)",
                       "Max Price/sqft (AED)","Total Sellable (sqft)","Total Value (AED)"]
        grp.to_excel(writer, index=False, sheet_name="Topology Summary")

        fe = []
        for fl in st.session_state.floors:
            ft = floor_total(fl, st.session_state.fm_params)
            for u in fl["units"]:
                v = unit_val(u["type"], u["rate"], st.session_state.fm_params)
                fe.append({"Floor": fl["floor"], "Kind": fl["kind"], "Levels": fl["levels"],
                           "Unit No": u["unit_no"], "Type": u["type"], "Rate (AED/sqft)": round(u["rate"],2),
                           "Unit Total (AED)": round(v["total"],0), "Floor Total (AED)": round(ft,0)})
        if fe:
            pd.DataFrame(fe).to_excel(writer, index=False, sheet_name="Floor Manager")

        for ws in writer.book.worksheets:        # apply professional styling to every sheet
            _style_export_sheet(ws)
    return out.getvalue()
# Note: the page-level "Download Updated Excel" button was removed — each table now has its
# own top-right "Export to Excel" button. build_export() is kept for potential reuse.
